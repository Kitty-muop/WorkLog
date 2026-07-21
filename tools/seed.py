import openpyxl
from openpyxl.styles import Font
from datetime import date, time

WB_PATH = r'D:\WorkLog\worklog.xlsx'
wb = openpyxl.load_workbook(WB_PATH)

EXAMPLE_DATE = date(2026, 7, 14)  # Tuesday
MAX_ROWS = 200

# ====== Clear old data ======

ws_te = wb['Time Entries']

# Unmerge all cells first
for m in list(ws_te.merged_cells.ranges):
    ws_te.unmerge_cells(str(m))

for r in range(2, MAX_ROWS + 1):
    for c in range(1, 18):
        if c not in (10, 12, 13, 14, 15, 16, 17):  # Keep J(10=Duration), L(12=Proj norm), M(13=Task norm), N(14=Combined), O(15=Unique ID), P(16=Date row), Q(17=Seq) — formula cols
            ws_te.cell(row=r, column=c).value = None

ws_kpi = wb['KPIs']
for r in range(2, MAX_ROWS + 1):
    for c in range(1, 10):
        ws_kpi.cell(row=r, column=c).value = None

# ====== KPIs example ======

ws_kpi.cell(row=2, column=1).value = 'KPI-1'
ws_kpi.cell(row=2, column=2).value = 'API Integration'
ws_kpi.cell(row=2, column=3).value = 'Backend'
ws_kpi.cell(row=2, column=4).value = EXAMPLE_DATE
ws_kpi.cell(row=2, column=5).value = 3
ws_kpi.cell(row=2, column=6).value = '=IF(E2<>"",E2*7.5,"")'  # formula from build
ws_kpi.cell(row=2, column=7).value = 'In Progress'
ws_kpi.cell(row=2, column=8).value = None

ws_kpi.cell(row=3, column=1).value = 'KPI-2'
ws_kpi.cell(row=3, column=2).value = 'Code Review'
ws_kpi.cell(row=3, column=3).value = 'Backend'
ws_kpi.cell(row=3, column=4).value = EXAMPLE_DATE
ws_kpi.cell(row=3, column=5).value = 1
ws_kpi.cell(row=3, column=6).value = '=IF(E3<>"",E3*7.5,"")'
ws_kpi.cell(row=3, column=7).value = 'Done'
ws_kpi.cell(row=3, column=8).value = EXAMPLE_DATE

# ====== Time Entries example ======

entries = [
    (EXAMPLE_DATE, 'Tue', 'Backend', 'API Integration', None, 'KPI-1-ST-1', 'Development',  time(9,0), time(11,0), None, 'Set up REST endpoints'),
    (EXAMPLE_DATE, 'Tue', 'Backend', 'API Integration', None, 'KPI-1-ST-2', 'Development',  time(11,15), time(12,30), None, 'Implement auth middleware'),
    (EXAMPLE_DATE, 'Tue', 'Backend', 'API Integration', None, 'KPI-1-ST-3', 'Testing',      time(13,30), time(15,0), None, 'Write integration tests'),
    (EXAMPLE_DATE, 'Tue', 'Backend', 'Code Review',     None, 'KPI-2-ST-1', 'Review',       time(15,15), time(16,0), None, 'Review PR #42'),
]

for i, (d, day, proj, task, sub, subcode, cat, start, end, dur, desc) in enumerate(entries):
    r = 2 + i
    ws_te.cell(row=r, column=1).value = d
    ws_te.cell(row=r, column=2).value = day
    ws_te.cell(row=r, column=3).value = proj
    ws_te.cell(row=r, column=4).value = task
    ws_te.cell(row=r, column=5).value = sub      # Sub Task
    ws_te.cell(row=r, column=6).value = subcode  # Sub Task Code
    ws_te.cell(row=r, column=7).value = cat
    ws_te.cell(row=r, column=8).value = start
    ws_te.cell(row=r, column=9).value = end
    ws_te.cell(row=r, column=11).value = desc

# ====== Projects example ======
ws_pr = wb['Projects']
ws_pr.cell(row=2, column=1).value = 'PRJ-1'
ws_pr.cell(row=2, column=2).value = 'Backend'
ws_pr.cell(row=2, column=3).value = 'API development and maintenance'
ws_pr.cell(row=2, column=4).value = 'Active'
ws_pr.cell(row=2, column=5).value = EXAMPLE_DATE

# Update KPI codes to project-scoped format
ws_kpi = wb['KPIs']
ws_kpi.cell(row=2, column=1).value = 'PRJ-1-KPI-1'
ws_kpi.cell(row=3, column=1).value = 'PRJ-1-KPI-2'

# Update subtask codes in Time Entries
ws_te.cell(row=2, column=6).value = 'PRJ-1-KPI-1-ST-1'
ws_te.cell(row=3, column=6).value = 'PRJ-1-KPI-1-ST-2'
ws_te.cell(row=4, column=6).value = 'PRJ-1-KPI-1-ST-3'
ws_te.cell(row=5, column=6).value = 'PRJ-1-KPI-2-ST-1'

wb.save(WB_PATH)
print("Seed data written.")
print(f"  Projects: 1 (PRJ-1: Backend)")
print(f"  KPIs: 2 (PRJ-1-KPI-1: API Integration, PRJ-1-KPI-2: Code Review)")
print(f"  Time Entries: {len(entries)} rows for {EXAMPLE_DATE}")

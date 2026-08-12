#!/usr/bin/env python3
"""WorkLog Static Dashboard Generator & GitHub Pages Deployer.

Reads worklog.xlsx and gamify data, generates dashboard/index.html,
and optionally deploys it to the gh-pages branch.
"""

import os
import sys
import json
import datetime
import subprocess
from pathlib import Path
from openpyxl import load_workbook

# Add project root to sys.path
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "tools"))

import gamify
from timer import WORKLOG_FILE

TEMPLATE_PATH = PROJECT_ROOT / "dashboard" / "template.html"
OUTPUT_PATH = PROJECT_ROOT / "dashboard" / "index.html"
GITHUB_PAGES_URL = "https://kitty-muop.github.io/WorkLog/"


def _safe_float(val, default=0.0):
    """Safely convert value to float with fallback."""
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _build_project_tree(wb, sub_actual_map):
    """Build hierarchical project -> KPI -> SubTask tree."""
    ws_p = wb["Projects"]
    ws_k = wb["KPIs"]
    ws_st = wb["SubTasks"]

    kpis_by_proj = {}
    for r in range(2, ws_k.max_row + 1):
        k_code = ws_k.cell(r, 1).value
        k_name = ws_k.cell(r, 2).value
        p_name = ws_k.cell(r, 3).value
        status = ws_k.cell(r, 7).value
        if k_name and p_name:
            kpis_by_proj.setdefault(str(p_name).strip(), []).append({
                "code": str(k_code or ""),
                "name": str(k_name).strip(),
                "status": str(status or "In Progress")
            })

    sub_by_kpi = {}
    for r in range(2, ws_st.max_row + 1):
        st_code = ws_st.cell(r, 1).value
        st_name = ws_st.cell(r, 2).value
        ft_name = ws_st.cell(r, 3).value
        status = ws_st.cell(r, 5).value
        est_val = ws_st.cell(r, 9).value
        if st_name and ft_name:
            st_str = str(st_name).strip()
            act_h = sub_actual_map.get(st_str, 0.0)
            est_h = _safe_float(est_val, default=7.5)

            sub_by_kpi.setdefault(str(ft_name).strip(), []).append({
                "code": str(st_code or ""),
                "name": st_str,
                "status": str(status or "In Progress"),
                "actual_h": act_h,
                "estimate_h": est_h
            })

    projects_tree = []
    for r in range(2, ws_p.max_row + 1):
        p_code = ws_p.cell(r, 1).value
        p_name = ws_p.cell(r, 2).value
        status = ws_p.cell(r, 4).value
        if not p_name:
            continue
        p_str = str(p_name).strip()

        proj_kpis = kpis_by_proj.get(p_str, [])
        for k in proj_kpis:
            k["subtasks"] = sub_by_kpi.get(k["name"], [])

        projects_tree.append({
            "code": str(p_code or ""),
            "name": p_str,
            "status": str(status or "In Progress"),
            "kpis": proj_kpis
        })

    return projects_tree


def extract_dashboard_data():
    """Extract all 5 sections of data from worklog.xlsx and gamify engine."""
    g_res = gamify.run()

    wb = load_workbook(WORKLOG_FILE, data_only=True)

    # 1. RPG Stats
    thresholds = gamify.SCORE_THRESHOLDS
    lvl = g_res.get("level", 0)
    cur_t = thresholds[lvl] if lvl < len(thresholds) else thresholds[-1]
    next_t = thresholds[lvl + 1] if lvl + 1 < len(thresholds) else cur_t * 2
    rpg_stats = {
        "hero_rank": g_res.get("level_name", "Novice"),
        "level": lvl,
        "total_exp": g_res.get("total_score", 0),
        "current_threshold": cur_t,
        "next_threshold": next_t,
        "streak": g_res.get("streak", 0),
        "max_streak": g_res.get("max_streak", 0),
        "consistency_pct": g_res.get("consistency_pct", 0.0),
        "today_score": g_res.get("today_score", 0),
    }

    # 2. Daily Hours (Last 14 Days) & Category Distribution
    ws_te = wb["Time Entries"]
    daily_map = {}
    cat_map = {}
    sub_actual_map = {}

    for r in range(2, ws_te.max_row + 1):
        d_val = ws_te.cell(r, 1).value
        st_val = ws_te.cell(r, 5).value
        cat_val = ws_te.cell(r, 7).value
        dur_val = ws_te.cell(r, 10).value

        dur = _safe_float(dur_val, default=0.0)
        if not d_val or dur <= 0:
            continue

        d_str = _cell_to_str(d_val)
        if d_str:
            daily_map[d_str] = round(daily_map.get(d_str, 0.0) + dur, 2)

        cat_str = str(cat_val or "Development").strip()
        cat_map[cat_str] = round(cat_map.get(cat_str, 0.0) + dur, 2)

        if st_val:
            st_key = str(st_val).strip()
            sub_actual_map[st_key] = round(sub_actual_map.get(st_key, 0.0) + dur, 2)

    # Sort daily hours and take last 14 days
    sorted_days = sorted(daily_map.items())[-14:]
    daily_hours = {
        "labels": [d[0] for d in sorted_days],
        "values": [d[1] for d in sorted_days],
        "target": 7.5
    }

    # 3. Category Distribution
    categories = {
        "labels": list(cat_map.keys()),
        "values": list(cat_map.values())
    }

    # 4. Estimate vs Actual Trend
    ws_st = wb["SubTasks"]
    est_act_labels = []
    estimates = []
    actuals = []

    for r in range(2, ws_st.max_row + 1):
        st_name = ws_st.cell(r, 2).value
        if not st_name or str(st_name).startswith("[Enter"):
            continue
        st_str = str(st_name).strip()
        est_h = _safe_float(ws_st.cell(r, 9).value, default=7.5)
        act_h = sub_actual_map.get(st_str, 0.0)

        est_act_labels.append(st_str[:25] + ("..." if len(st_str) > 25 else ""))
        estimates.append(est_h)
        actuals.append(act_h)

    estimate_vs_actual = {
        "labels": est_act_labels[-10:],
        "estimates": estimates[-10:],
        "actuals": actuals[-10:]
    }

    # 5. Project Tree
    projects_tree = _build_project_tree(wb, sub_actual_map)

    return {
        "generated_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S ICT"),
        "rpg_stats": rpg_stats,
        "daily_hours": daily_hours,
        "categories": categories,
        "estimate_vs_actual": estimate_vs_actual,
        "projects_tree": projects_tree
    }

    return {
        "generated_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S ICT"),
        "rpg_stats": rpg_stats,
        "daily_hours": daily_hours,
        "categories": categories,
        "estimate_vs_actual": estimate_vs_actual,
        "projects_tree": projects_tree
    }


def generate_html():
    """Generate index.html from template.html with embedded json data."""
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Template not found at {TEMPLATE_PATH}")

    data = extract_dashboard_data()
    template_content = TEMPLATE_PATH.read_text(encoding="utf-8")

    json_str = json.dumps(data, indent=2)
    html_content = template_content.replace("__DATA_PLACEHOLDER__", json_str)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(html_content, encoding="utf-8")
    print(f"[DASHBOARD] Generated HTML at {OUTPUT_PATH}")
    return OUTPUT_PATH


def deploy_dashboard():
    """Generate index.html and push to gh-pages branch on GitHub via isolated worktree."""
    generate_html()

    cwd = str(PROJECT_ROOT)
    tmp_dir = PROJECT_ROOT / ".gh_pages_build"

    try:
        # Fetch gh-pages branch from remote if available
        subprocess.run(["git", "fetch", "origin", "gh-pages"], capture_output=True, cwd=cwd)

        # Remove stale worktree if exists
        if tmp_dir.exists():
            subprocess.run(["git", "worktree", "remove", "-f", str(tmp_dir)], capture_output=True, cwd=cwd)

        # Create isolated worktree for gh-pages branch
        res = subprocess.run(["git", "worktree", "add", "-B", "gh-pages", str(tmp_dir)], capture_output=True, cwd=cwd)
        if res.returncode != 0:
            # Create orphan worktree if gh-pages branch does not exist yet
            subprocess.run(["git", "worktree", "add", "--detach", str(tmp_dir)], check=True, cwd=cwd)
            subprocess.run(["git", "checkout", "--orphan", "gh-pages"], capture_output=True, cwd=str(tmp_dir))
            subprocess.run(["git", "rm", "-rf", "."], capture_output=True, cwd=str(tmp_dir))

        # Copy generated index.html into the isolated worktree
        (tmp_dir / "index.html").write_text(OUTPUT_PATH.read_text(encoding="utf-8"), encoding="utf-8")

        # Commit and push from the isolated worktree directory
        subprocess.run(["git", "add", "index.html"], check=True, cwd=str(tmp_dir))
        subprocess.run(["git", "commit", "-m", f"chore: update dashboard static site ({datetime.datetime.now().strftime('%Y-%m-%d %H:%M')})"], capture_output=True, cwd=str(tmp_dir))
        subprocess.run(["git", "push", "origin", "gh-pages", "--force"], check=True, cwd=str(tmp_dir))

        # Remove isolated worktree
        subprocess.run(["git", "worktree", "remove", "-f", str(tmp_dir)], capture_output=True, cwd=cwd)

        print(f"[DASHBOARD] Deployed successfully to GitHub Pages: {GITHUB_PAGES_URL}")
        return True, GITHUB_PAGES_URL
    except Exception as e:
        print(f"[DASHBOARD] Deploy error: {e}")
        if tmp_dir.exists():
            subprocess.run(["git", "worktree", "remove", "-f", str(tmp_dir)], capture_output=True, cwd=cwd)
        return False, str(e)


if __name__ == "__main__":
    generate_html()

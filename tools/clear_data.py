#!/usr/bin/env python3
"""Clear all data rows in worklog.xlsx, keeping headers and formulas.
Then sync the cleared state to Google Sheets."""

import os
import sys
from pathlib import Path

import openpyxl

WB_PATH = Path(__file__).parent.parent / 'worklog.xlsx'
MAX_DATA_ROWS = 500


def clear_xlsx():
    wb = openpyxl.load_workbook(WB_PATH)

    # ---- Time Entries ----
    ws = wb['Time Entries']
    for r in range(2, MAX_DATA_ROWS + 2):
        for c in range(1, 13):
            ws.cell(row=r, column=c).value = None

    # ---- Projects ----
    ws = wb['Projects']
    for r in range(2, MAX_DATA_ROWS + 2):
        for c in range(1, 6):
            ws.cell(row=r, column=c).value = None

    # ---- KPIs ----
    ws = wb['KPIs']
    for r in range(2, MAX_DATA_ROWS + 2):
        for c in [1, 2, 3, 4, 5, 7, 8, 9]:
            ws.cell(row=r, column=c).value = None

    # ---- SubTasks ----
    ws = wb['SubTasks']
    for r in range(2, MAX_DATA_ROWS + 2):
        for c in range(1, 9):
            ws.cell(row=r, column=c).value = None

    wb.save(WB_PATH)
    print(f"Cleared data from {WB_PATH}")


def clear_gsheets():
    """Clear all data rows in Google Sheets for Time Entries, Projects, KPIs."""
    import gspread
    from google.oauth2.service_account import Credentials

    sheet_id = os.environ.get('WORKLOG_GSHEETS_ID')
    key_path = os.environ.get('WORKLOG_GSHEETS_KEY')

    if not sheet_id or not key_path:
        print("Google Sheets not configured (WORKLOG_GSHEETS_ID / WORKLOG_GSHEETS_KEY not set). Skipping.")
        return False

    key_path = Path(key_path).expanduser()
    if not key_path.exists():
        print(f"Service account key not found at {key_path}. Skipping.")
        return False

    creds = Credentials.from_service_account_file(
        str(key_path),
        scopes=['https://www.googleapis.com/auth/spreadsheets']
    )
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(sheet_id)

    for sheet_name in ['Time Entries', 'Projects', 'KPIs', 'SubTasks']:
        try:
            ws = sh.worksheet(sheet_name)
            all_vals = ws.get_all_values()
            if len(all_vals) > 1:
                # Keep header row, clear everything below
                ws.batch_clear([f'A2:{chr(64 + ws.col_count)}{len(all_vals)}'])
                print(f"  Cleared Google Sheet '{sheet_name}'")
        except Exception as e:
            print(f"  Could not clear '{sheet_name}': {e}")

    print("Google Sheets cleared successfully.")
    return True


if __name__ == '__main__':
    clear_xlsx()
    clear_gsheets()

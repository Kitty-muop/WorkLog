import json
import openpyxl
import pandas as pd
from datetime import date, datetime
from pathlib import Path

WORKLOG = Path("D:/WorkLog/worklog.xlsx")
SCORE_FILE = Path("D:/WorkLog/.gamify_data.json")

PRODUCTION_CATEGORIES = [
    "Development",
    "Debug / Bug Fix",
    "Refactoring",
    "Code Review",
    "Testing / QA",
    "DevOps / CI-CD",
    "Documentation"
]

def generate_level_thresholds():
    """Generate 101 level thresholds (Level 0 to Level 100) across 5 major tiers."""
    thresholds = [0]
    curr = 0
    for lvl in range(1, 101):
        if lvl <= 20:
            step = 50       # Tier 1: Novice (0 - 1,000 EXP)
        elif lvl <= 40:
            step = 100      # Tier 2: Adventurer (1,001 - 3,000 EXP)
        elif lvl <= 60:
            step = 200      # Tier 3: Expert (3,001 - 7,000 EXP)
        elif lvl <= 80:
            step = 400      # Tier 4: Master (7,001 - 15,000 EXP)
        else:
            step = 800      # Tier 5: Grandmaster Legend (15,001 - 31,000 EXP)
        curr += step
        thresholds.append(curr)
    return thresholds

SCORE_THRESHOLDS = generate_level_thresholds()

def get_tier_name(level):
    """Map Level (0..100) to 5 Major Level Milestones."""
    if level <= 20:
        return "Novice"
    elif level <= 40:
        return "Adventurer"
    elif level <= 60:
        return "Expert"
    elif level <= 80:
        return "Master"
    else:
        return "Grandmaster Legend"

def calculate_task_exp(estimated_h, actual_h, category="Development"):
    """Calculate EXP based on difficulty/estimate, actual/estimate ratio, and category bonuses."""
    est = float(estimated_h or 1.0)
    act = float(actual_h or 1.0)
    cat_lower = str(category or "").strip().lower()
    is_debug = any(k in cat_lower for k in ("debug", "bug", "fix", "troubleshoot"))

    # Base EXP proportional to estimated effort (20 EXP per estimated hour)
    base_exp = est * 20.0

    if act <= est:
        # Efficiency Bonus for completing on time or early
        ratio = act / est if est > 0 else 1.0
        efficiency_factor = 1.0 + (1.0 - ratio) * 0.5
        exp = base_exp * efficiency_factor

        # Debug category on-time bonus (+30% EXP)
        if is_debug:
            exp *= 1.30
    else:
        # Penalty for exceeding estimated time
        penalty_ratio = est / act if act > 0 else 0.5

        if is_debug:
            # Heavy penalty for debug/bug fix exceeding estimate
            penalty_factor = max(0.05, penalty_ratio ** 2)
        else:
            penalty_factor = max(0.2, penalty_ratio)

        exp = base_exp * penalty_factor

    return round(exp, 1), is_debug

DAILY_TARGET_HOURS = 7.5

def load_gamify_data():
    if SCORE_FILE.exists():
        return json.loads(SCORE_FILE.read_text())
    return {
        "total_score": 0,
        "streak": 0,
        "max_streak": 0,
        "last_logged_date": None,
        "daily_history": {},
        "level": 0,
    }

def save_gamify_data(data):
    SCORE_FILE.write_text(json.dumps(data, indent=2, default=str))

def get_dur(val):
    return float(val) if isinstance(val, (int, float)) else 0.0


def get_time_entries(ws):
    entries = []
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        date_raw = r[0]
        project = r[2]
        task = r[3]
        category = r[6]
        dur = get_dur(r[9])
        desc = r[10] if len(r) > 10 else ''
        if date_raw and dur > 0:
            entries.append({
                "date": pd.Timestamp(date_raw).date(),
                "project": str(project or ""),
                "task": str(task or ""),
                "category": str(category or ""),
                "duration": dur,
                "description": str(desc or ""),
            })
    return entries

def compute_daily_summary(entries):
    daily = {}
    for e in entries:
        d = e["date"]
        if d not in daily:
            daily[d] = {"total_hours": 0, "task_count": 0, "categories": set()}
        daily[d]["total_hours"] += e["duration"]
        daily[d]["task_count"] += 1
        daily[d]["categories"].add(e["category"])
    return daily

def update_streak(data, daily_summary, today):
    sorted_dates = sorted(daily_summary.keys())
    if not sorted_dates:
        data["streak"] = 0
        data["last_logged_date"] = None
        return

    # Check if logged today
    last = sorted_dates[-1]
    data["last_logged_date"] = str(last)

    # Recalculate streak from today backwards
    streak = 0
    check = today
    while check in daily_summary or check.weekday() >= 5:
        if check.weekday() < 5 and check not in daily_summary:
            break
        if check in daily_summary:
            streak += 1
        check -= pd.Timedelta(days=1)
    data["streak"] = streak
    if streak > data["max_streak"]:
        data["max_streak"] = streak

def compute_completion_bonus(wb):
    """Calculate bonus EXP for completed Subtasks, KPIs, and Projects."""
    completion_exp = 0
    counts = {"subtasks": 0, "kpis": 0, "projects": 0}

    # 1. Subtasks (+15 EXP each)
    if 'SubTasks' in wb.sheetnames:
        ws_sub = wb['SubTasks']
        for r in ws_sub.iter_rows(min_row=2, values_only=True):
            st = str(r[4] or '').lower() if len(r) > 4 else ''
            if st in ('done', 'completed'):
                completion_exp += 15
                counts["subtasks"] += 1

    # 2. KPIs (+50 EXP each)
    if 'KPIs' in wb.sheetnames:
        ws_kpi = wb['KPIs']
        for r in ws_kpi.iter_rows(min_row=2, values_only=True):
            st = str(r[6] or '').lower() if len(r) > 6 else ''
            if st in ('done', 'completed'):
                completion_exp += 50
                counts["kpis"] += 1

    # 3. Projects (+150 EXP each)
    if 'Projects' in wb.sheetnames:
        ws_proj = wb['Projects']
        for r in ws_proj.iter_rows(min_row=2, values_only=True):
            st = str(r[3] or '').lower() if len(r) > 3 else ''
            if st in ('done', 'completed'):
                completion_exp += 150
                counts["projects"] += 1

    return completion_exp, counts


def compute_scores(entries, daily_summary, data, today, wb=None):
    today_str = str(today)
    today_entry = daily_summary.get(today, None)
    today_score = 0

    if today_entry:
        util_pct = min(today_entry["total_hours"] / DAILY_TARGET_HOURS, 1.0)
        today_score += round(util_pct * 50)

        task_bonus = min(today_entry["task_count"] * 5, 15)
        today_score += task_bonus

        category_count = len(today_entry["categories"])
        if category_count >= 2:
            today_score += 5

    streak_bonus = min(data["streak"] * 2, 20)
    today_score += streak_bonus

    daily_history = data.get("daily_history", {})
    daily_history[today_str] = {
        "date": today_str,
        "score": today_score,
        "hours": today_entry["total_hours"] if today_entry else 0,
        "tasks": today_entry["task_count"] if today_entry else 0,
    }
    data["daily_history"] = daily_history

    # Sum all historical daily scores
    work_score = sum(day.get('score', 0) for day in daily_history.values() if isinstance(day, dict))
    
    # Add completion bonus
    completion_exp = 0
    counts = {"subtasks": 0, "kpis": 0, "projects": 0}
    if wb:
        completion_exp, counts = compute_completion_bonus(wb)
    
    data["total_score"] = work_score + completion_exp
    data["completion_exp"] = completion_exp
    data["completed_counts"] = counts

    # Determine level
    for i, threshold in reversed(list(enumerate(SCORE_THRESHOLDS))):
        if data["total_score"] >= threshold:
            data["level"] = i
            break

    return today_score

def compute_consistency(entries):
    if not entries:
        return 0.0, 0, 0
    dates = sorted(set(e["date"] for e in entries))
    if not dates:
        return 0.0, 0, 0
    mn, mx = dates[0], dates[-1]
    total_weekdays = sum(
        1 for d in pd.date_range(mn, mx, freq='D')
        if d.weekday() < 5
    )
    logged_weekdays = sum(1 for d in dates if d.weekday() < 5)
    pct = (logged_weekdays / total_weekdays * 100) if total_weekdays > 0 else 0
    return round(pct, 1), logged_weekdays, total_weekdays

def compute_task_performance(ws, entries):
    ws_kpi = openpyxl.load_workbook(WORKLOG, data_only=True)['KPIs']
    tasks = {}
    for r in ws_kpi.iter_rows(min_row=2, max_row=ws_kpi.max_row, values_only=True):
        code, name, project, _, days, hours_deadline, status = r[0], r[1], r[2], r[3], r[4], r[5], r[6]
        if name:
            deadline_h = hours_deadline if hours_deadline else (days * 7.5 if days else None)
            logged = sum(e["duration"] for e in entries if e["task"] == str(name))
            pct = (logged / deadline_h * 100) if deadline_h and deadline_h > 0 else 0
            tasks[str(name)] = {
                "code": str(code or ''),
                "project": str(project),
                "logged_hours": round(logged, 2),
                "deadline_hours": deadline_h,
                "performance_pct": round(pct, 1),
                "status": str(status),
            }
    return tasks

def run():
    wb = openpyxl.load_workbook(WORKLOG, data_only=True)
    ws = wb['Time Entries']
    today = date.today()

    entries = get_time_entries(ws)
    daily = compute_daily_summary(entries)

    data = load_gamify_data()
    update_streak(data, daily, today)
    today_score = compute_scores(entries, daily, data, today, wb=wb)
    consistency, logged_days, total_days = compute_consistency(entries)
    task_perf = compute_task_performance(ws, entries)

    save_gamify_data(data)

    return {
        "today_score": today_score,
        "total_score": data["total_score"],
        "level": data["level"],
        "level_name": get_tier_name(data["level"]),
        "streak": data["streak"],
        "max_streak": data["max_streak"],
        "consistency_pct": consistency,
        "logged_weekdays": logged_days,
        "total_weekdays": total_days,
        "tasks": task_perf,
        "daily_summary": {
            str(k): {
                "hours": round(v["total_hours"], 2),
                "tasks": v["task_count"],
                "categories": list(v["categories"]),
            }
            for k, v in daily.items()
        },
    }

if __name__ == '__main__':
    result = run()
    print(json.dumps(result, indent=2, default=str))

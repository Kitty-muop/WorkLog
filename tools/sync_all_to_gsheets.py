#!/usr/bin/env python3
"""Sync ALL existing data from worklog.xlsx to Google Sheets."""

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))

import openpyxl
import gsheets

WB = Path(__file__).parent.parent / 'worklog.xlsx'
MAX = 500

def clear_gsheet(ws, num_cols):
    """Clear all data rows, keep header."""
    try:
        all_vals = ws.get_all_values()
        if len(all_vals) > 1:
            ws.batch_clear([f'A2:{chr(64 + num_cols)}{len(all_vals)}'])
    except:
        pass

def find_last_data_row(ws, key_col):
    """Find last row with data in key column."""
    for r in range(MAX + 1, 1, -1):
        val = ws.cell(r, key_col).value
        if val is not None:
            return r
    return 1


def sync_sheet(sheet_name, col_map, key_col=2):
    """Copy rows from Excel to Google Sheets."""
    from google.oauth2.service_account import Credentials
    import gspread
    import os

    key_path = Path(os.environ.get('WORKLOG_GSHEETS_KEY', ''))
    sheet_id = os.environ.get('WORKLOG_GSHEETS_ID', '')
    if not key_path.exists() or not sheet_id:
        print(f"  SKIP {sheet_name}: env not configured")
        return

    creds = Credentials.from_service_account_file(str(key_path), scopes=['https://www.googleapis.com/auth/spreadsheets'])
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(sheet_id)

    try:
        gws = sh.worksheet(sheet_name)
    except:
        print(f"  SKIP {sheet_name}: sheet not found in Google Sheets")
        return

    clear_gsheet(gws, max(col_map.values()))

    wb = openpyxl.load_workbook(WB)
    ws = wb[sheet_name]
    last_row = find_last_data_row(ws, key_col)
    rows = []
    for r in range(2, last_row + 1):
        row_data = []
        all_empty = True
        for col_name, col_idx in sorted(col_map.items(), key=lambda x: x[1]):
            cell_val = ws.cell(r, col_idx).value
            if cell_val is not None:
                all_empty = False
            # Convert datetime/date to string
            if hasattr(cell_val, 'strftime'):
                if hasattr(cell_val, 'hour'):  # datetime
                    cell_val = cell_val.strftime('%d-%m-%Y %H:%M')
                else:  # date
                    cell_val = cell_val.strftime('%d-%m-%Y')
            elif isinstance(cell_val, float) and col_name == 'Duration (h)':
                cell_val = round(cell_val, 2)
            row_data.append(str(cell_val) if cell_val is not None else '')
        if all_empty:
            continue
        rows.append(row_data)

    if rows:
        range_str = f'A2:{chr(64 + max(col_map.values()))}{len(rows) + 1}'
        gws.update(range_name=range_str, values=rows, value_input_option='USER_ENTERED')
        print(f"  {sheet_name}: synced {len(rows)} rows")
    else:
        print(f"  {sheet_name}: no data to sync")


print("Syncing all data to Google Sheets...\n")

# Projects: Code(1), Name(2), Description(3), Status(4), Created Date(5)
sync_sheet('Projects', {'Code': 1, 'Project Name': 2, 'Description': 3, 'Status': 4, 'Created Date': 5})

# KPIs: Code(1), Father Task(2), Project(3), Date(4), Deadline Days(5), Deadline H(6), Status(7), Completed Date(8)
sync_sheet('KPIs', {
    'Code': 1, 'Father Task': 2, 'Project': 3, 'Date': 4,
    'Deadline (days)': 5, 'Deadline (h)': 6, 'Status': 7, 'Completed Date': 8
})

# SubTasks: Code(1), Name(2), Father Task(3), Project(4), Status(5), Created Date(6), Completed Date(7)
sync_sheet('SubTasks', {
    'Code': 1, 'Sub Task': 2, 'Father Task': 3, 'Project': 4,
    'Status': 5, 'Created Date': 6, 'Completed Date': 7
})

# Time Entries: full 12 cols (key_col=1 for Date)
sync_sheet('Time Entries', {
    'Date': 1, 'Day': 2, 'Project': 3, 'Father Task': 4, 'Sub Task': 5,
    'Sub Task Code': 6, 'Category': 7, 'Start Time': 8, 'End Time': 9,
    'Duration (h)': 10, 'Description': 11, 'Break Info': 12
}, key_col=1)

print("\nDone!")

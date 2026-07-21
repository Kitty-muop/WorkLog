import openpyxl
import pandas as pd
from pathlib import Path

WORKLOG = Path("D:/WorkLog/worklog.xlsx")

def load_ws():
    return openpyxl.load_workbook(WORKLOG, data_only=True)['Time Entries']

def empty_rows(ws):
    count = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=10):
        if all(c.value is None for c in row):
            count += 1
    return count

def get_dur(val):
    """Get numeric duration from cell value (formula, number, or None)."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    if hasattr(val, 'hour'):  # datetime.time from formula-less open
        return 0.0
    return 0.0


def missing_descriptions(ws):
    out = []
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        date, dur, desc = r[0], r[9], r[10]
        dur = get_dur(dur)
        if date and dur > 0 and (desc is None or str(desc).strip() in ("", "0")):
            out.append({"date": date, "duration": dur})
    return out

def missing_categories(ws):
    out = []
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        date, cat, dur = r[0], r[6], r[9]
        dur = get_dur(dur)
        if date and dur > 0 and (cat is None or str(cat).strip() == ""):
            out.append({"date": date, "duration": dur})
    return out

def missing_projects(ws):
    """Check that all projects in Time Entries exist in Projects sheet."""
    try:
        wb = openpyxl.load_workbook(WORKLOG, data_only=True)
        ws_pr = wb['Projects']
        known = set()
        for r in ws_pr.iter_rows(min_row=2, max_row=ws_pr.max_row, values_only=True):
            name = r[1]
            if name:
                known.add(str(name).strip().lower())
        missing = set()
        for r in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            proj = r[2]
            if proj and str(proj).strip().lower() not in known:
                missing.add(str(proj).strip())
        return sorted(missing)
    except Exception:
        return []


def utilization_gaps(ws):
    daily = {}
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        date, dur = r[0], r[9]
        dur = get_dur(dur)
        if date and dur > 0:
            d = pd.Timestamp(date).date()
            daily[d] = daily.get(d, 0) + dur
    if not daily:
        return []
    mn, mx = min(daily), max(daily)
    gaps = []
    for d in pd.date_range(mn, mx, freq='D'):
        if d.weekday() < 5 and d.date() not in daily:
            gaps.append(d.date())
    return gaps

def run():
    ws = load_ws()
    issues = []

    e = empty_rows(ws)
    if e:
        issues.append({"type": "EMPTY_ROWS", "severity": "medium", "detail": f"{e} empty rows"})

    for m in missing_descriptions(ws):
        issues.append({"type": "MISSING_DESC", "severity": "low", "detail": f"{m['date']}: {m['duration']}h no description"})

    for m in missing_categories(ws):
        issues.append({"type": "MISSING_CATEGORY", "severity": "low", "detail": f"{m['date']}: {m['duration']}h no category"})

    for g in utilization_gaps(ws):
        issues.append({"type": "UTILIZATION_GAP", "severity": "medium", "detail": f"No hours on {g} (weekday)"})

    for p in missing_projects(ws):
        issues.append({"type": "MISSING_PROJECT", "severity": "high", "detail": f"Project '{p}' not registered in Projects sheet"})

    return issues

if __name__ == '__main__':
    for i in run():
        print(f"[{i['severity'].upper()}] {i['type']}: {i['detail']}")

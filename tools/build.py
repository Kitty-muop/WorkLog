import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy
import datetime

MAX_DATA_ROWS = 500
NUM_WEEKS = 1
DD_ROWS = 60
WS_ROWS = 60
MO_ROWS = 50

wb = openpyxl.Workbook()

LIGHT_BLUE_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
LIGHT_GRAY_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
NORMAL_FONT = Font(name='Calibri', size=11)
BOLD_FONT = Font(name='Calibri', size=11, bold=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

# ===================================================================
# SHEET 1: Time Entries
# ===================================================================
ws_te = wb.active
ws_te.title = 'Time Entries'

ws_te.column_dimensions['A'].width = 13
ws_te.column_dimensions['B'].width = 5
ws_te.column_dimensions['C'].width = 15
ws_te.column_dimensions['D'].width = 25
ws_te.column_dimensions['E'].width = 20
ws_te.column_dimensions['F'].width = 15
ws_te.column_dimensions['G'].width = 15
ws_te.column_dimensions['H'].width = 11
ws_te.column_dimensions['I'].width = 11
ws_te.column_dimensions['J'].width = 12
ws_te.column_dimensions['K'].width = 50
ws_te.column_dimensions['L'].width = 30
ws_te.column_dimensions['M'].width = 15
ws_te.column_dimensions['N'].width = 25
ws_te.column_dimensions['O'].width = 30
ws_te.column_dimensions['P'].width = 8
ws_te.column_dimensions['Q'].width = 8
ws_te.column_dimensions['R'].width = 8

for c_letter in ['M', 'N', 'O', 'P', 'Q', 'R']:
    ws_te.column_dimensions[c_letter].hidden = True

headers_te = ['Date', 'Day', 'Project', 'Father Task', 'Sub Task', 'Sub Task Code', 'Category',
              'Start Time', 'End Time', 'Duration (h)', 'Description', 'Break Info']
for i, h in enumerate(headers_te, 1):
    ws_te.cell(row=1, column=i, value=h)
style_header(ws_te, 1, 12)

# Helper column formulas (rows 2..MAX_DATA_ROWS+1)
for r in range(2, MAX_DATA_ROWS + 2):
    prev_row = r - 1
    # M: Project norm
    if r == 2:
        ws_te.cell(row=r, column=13).value = '=IF(C2<>"",C2,"")'
    else:
        ws_te.cell(row=r, column=13).value = f'=IF(C{r}<>"",C{r},M{prev_row})'

    # N: Father Task norm
    if r == 2:
        ws_te.cell(row=r, column=14).value = '=IF(D2<>"",D2,"")'
    else:
        ws_te.cell(row=r, column=14).value = f'=IF(D{r}<>"",D{r},N{prev_row})'

    # O: Project:Task combined
    ws_te.cell(row=r, column=15).value = f'=M{r}&": "&N{r}'

    # P: Unique task ID (for Weekly Summary)
    if r == 2:
        ws_te.cell(row=r, column=16).value = (
            '=IF(O2="","",IF(COUNTIF($O$1:O1,O2)=0,MAX($P$1:P1)+1,""))'
        )
    else:
        ws_te.cell(row=r, column=16).value = (
            f'=IF(O{r}="","",IF(COUNTIF($O$1:O{prev_row},O{r})=0,MAX($P$1:P{prev_row})+1,""))'
        )

    # Q: Date row number (for Daily Detail)
    ws_te.cell(row=r, column=17).value = (
        f'=IF(A{r}="","",COUNTIF($A$2:A{r},A{r}))'
    )

    # R: Sequential entry number
    if r == 2:
        ws_te.cell(row=r, column=18).value = '=IF(A2="","",1)'
    else:
        ws_te.cell(row=r, column=18).value = f'=IF(A{r}="","",MAX($R$1:R{prev_row})+1)'

# Duration formula (col J)
for r in range(2, MAX_DATA_ROWS + 2):
    ws_te.cell(row=r, column=10).value = (
        f'=IF(AND(H{r}<>"",I{r}<>""),(I{r}-H{r})*24,"")'
    )
    ws_te.cell(row=r, column=10).number_format = '0.00'

# Date format A, time format H-I
for r in range(2, MAX_DATA_ROWS + 2):
    ws_te.cell(row=r, column=1).number_format = 'dd-mm-yyyy'
    ws_te.cell(row=r, column=8).number_format = 'hh:mm'
    ws_te.cell(row=r, column=9).number_format = 'hh:mm'

# Data validation for Status (col F) — not applicable here, but keep for KPIs
ws_te.freeze_panes = 'A2'
ws_te.auto_filter.ref = f'A1:L{MAX_DATA_ROWS + 1}'

print("Sheet 1 done: Time Entries with WPS-compatible helpers")

# ===================================================================
# SHEET 2: KPIs
# ===================================================================
ws_kp = wb.create_sheet('KPIs', 1)

ws_kp.column_dimensions['A'].width = 10
ws_kp.column_dimensions['B'].width = 30
ws_kp.column_dimensions['C'].width = 20
ws_kp.column_dimensions['D'].width = 15
ws_kp.column_dimensions['E'].width = 16
ws_kp.column_dimensions['F'].width = 14
ws_kp.column_dimensions['G'].width = 16
ws_kp.column_dimensions['H'].width = 16
ws_kp.column_dimensions['I'].width = 40

headers_kp = ['Code', 'Father Task', 'Project', 'Date', 'Deadline (days)',
              'Deadline (h)', 'Status', 'Completed Date', 'Notes']
for i, h in enumerate(headers_kp, 1):
    ws_kp.cell(row=1, column=i, value=h)
style_header(ws_kp, 1, 9)

# Deadline(h) formula
for r in range(2, MAX_DATA_ROWS + 2):
    ws_kp.cell(row=r, column=6).value = f'=IF(E{r}<>"",E{r}*7.5,"")'
    ws_kp.cell(row=r, column=6).number_format = '0.0'

# Date format
for r in range(2, MAX_DATA_ROWS + 2):
    ws_kp.cell(row=r, column=4).number_format = 'dd-mm-yyyy'
    ws_kp.cell(row=r, column=8).number_format = 'dd-mm-yyyy'

# Data validation (Status dropdown)
dv_status = DataValidation(
    type='list',
    formula1='"Not Started,In Progress,Done,Late"',
    allow_blank=True
)
dv_status.error = 'Please select a valid status'
dv_status.errorTitle = 'Invalid Status'
ws_kp.add_data_validation(dv_status)
dv_status.add(f'G2:G{MAX_DATA_ROWS + 1}')

# Sample header placeholder
ws_kp.cell(row=2, column=2).value = '[Enter father task name]'
ws_kp.cell(row=2, column=2).font = Font(name='Calibri', size=11, italic=True, color='999999')

ws_kp.freeze_panes = 'A2'
print("Sheet 2 done: KPIs")

# ===================================================================
# SHEET 3: Weekly Summary
# ===================================================================
ws_ws = wb.create_sheet('Weekly Summary', 2)

ws_ws.column_dimensions['A'].width = 45
ws_ws.column_dimensions['B'].width = 15
for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I']:
    ws_ws.column_dimensions[col].width = 14
ws_ws.column_dimensions['J'].width = 12

# Hidden ref cells
ws_ws.cell(row=1, column=10).value = 'LatestDate'
ws_ws.cell(row=2, column=10).value = "=MAX('Time Entries'!A:A)"
ws_ws.cell(row=1, column=11).value = 'WeekStart'
ws_ws.cell(row=2, column=11).value = '=J2-WEEKDAY(J2,3)'
ws_ws.cell(row=1, column=12).value = 'WeekEnd'
ws_ws.cell(row=2, column=12).value = '=J2+6'
ws_ws.column_dimensions['J'].hidden = True
ws_ws.column_dimensions['K'].hidden = True
ws_ws.column_dimensions['L'].hidden = True

# Row 2: Week label
ws_ws.cell(row=2, column=1).value = '="Week: "&TEXT(K2,"dd-mm-yyyy")&" - "&TEXT(L2,"dd-mm-yyyy")'
ws_ws.cell(row=2, column=1).font = Font(name='Calibri', size=12, bold=True, italic=True)
ws_ws.merge_cells('A2:I2')

# Row 4: Day headers (Mon-Sun)
ws_ws.cell(row=4, column=1).value = 'Task'
ws_ws.cell(row=4, column=2).value = 'Category'
weekdays = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
for i, wd in enumerate(weekdays):
    col = 3 + i
    ws_ws.cell(row=4, column=col).value = f'={get_column_letter(col-3+10)}2+{i}'
    ws_ws.cell(row=4, column=col).number_format = 'dd-mm-yyyy'
    ws_ws.cell(row=4, column=col).alignment = CENTER
ws_ws.cell(row=4, column=10).value = 'Total'
style_header(ws_ws, 4, 10)

# Helper col: Store date references (hidden, after col I)
# Cols O-U store actual date values for each weekday
for i in range(7):
    cell = ws_ws.cell(row=4, column=15 + i)
    cell.value = f'=K2+{i}'
    cell.number_format = '0'

ws_start = 5

# Rows 5+: Pre-populated tasks
for r in range(ws_start, ws_start + WS_ROWS):
    row_offset = r - ws_start + 1
    # A: Task name (from unique task list)
    ws_ws.cell(row=r, column=1).value = (
        f"=IFERROR(INDEX('Time Entries'!$O:$O,MATCH(ROW()-{ws_start-1},"
        f"'Time Entries'!$P:$P,0)),\"\")"
    )
    # B: Category
    ws_ws.cell(row=r, column=2).value = (
        f'=IF($A{r}="","",IFERROR(INDEX(\'Time Entries\'!G:G,'
        f'MATCH($A{r},\'Time Entries\'!$O:$O,0)),""))'
    )
    # C-I: Hours per day
    for i in range(7):
        col = 3 + i
        date_ref_col = get_column_letter(15 + i)
        ws_ws.cell(row=r, column=col).value = (
            f'=IF($A{r}="","",SUMIFS(\'Time Entries\'!J:J,\'Time Entries\'!O:O,$A{r},'
            f"\'Time Entries\'!A:A,$O$4+{i}))"
        )
        ws_ws.cell(row=r, column=col).number_format = '0.00'
    # J: Total
    ws_ws.cell(row=r, column=10).value = f'=IF($A{r}="","",SUM(C{r}:I{r}))'
    ws_ws.cell(row=r, column=10).number_format = '0.00'

# Grand Total row
gt_row = ws_start + WS_ROWS + 1
ws_ws.cell(row=gt_row, column=1).value = 'GRAND TOTAL'
ws_ws.cell(row=gt_row, column=1).font = BOLD_FONT
for i in range(9):
    col = 3 + i
    ws_ws.cell(row=gt_row, column=col).value = f'=SUM({get_column_letter(col)}{ws_start}:{get_column_letter(col)}{ws_start+WS_ROWS-1})'
    ws_ws.cell(row=gt_row, column=col).number_format = '0.00'
    ws_ws.cell(row=gt_row, column=col).font = BOLD_FONT

ws_ws.freeze_panes = 'A5'
print("Sheet 3 done: Weekly Summary (WPS-compatible)")

# ===================================================================
# SHEET 4: Daily Detail
# ===================================================================
ws_dd = wb.create_sheet('Daily Detail', 3)

ws_dd.column_dimensions['A'].width = 18
ws_dd.column_dimensions['B'].width = 20
ws_dd.column_dimensions['C'].width = 30
ws_dd.column_dimensions['D'].width = 15
ws_dd.column_dimensions['E'].width = 15
ws_dd.column_dimensions['F'].width = 12
ws_dd.column_dimensions['G'].width = 50
ws_dd.column_dimensions['H'].width = 10
ws_dd.column_dimensions['I'].width = 10

# Hide helper cols H and I
ws_dd.column_dimensions['H'].hidden = True
ws_dd.column_dimensions['I'].hidden = True

# Row 1: Title + date selector
ws_dd.cell(row=1, column=1).value = 'Daily Detail'
ws_dd.cell(row=1, column=1).font = Font(name='Calibri', size=14, bold=True)
ws_dd.cell(row=1, column=3).value = "=MAX('Time Entries'!A:A)"
ws_dd.cell(row=1, column=3).number_format = 'dd-mm-yyyy'
ws_dd.cell(row=1, column=3).font = Font(name='Calibri', size=12)
ws_dd.cell(row=1, column=3).fill = YELLOW_FILL
ws_dd.cell(row=1, column=2).value = 'Date:'
ws_dd.cell(row=1, column=2).font = BOLD_FONT
ws_dd.cell(row=1, column=4).value = '(Edit C1 to change date)'
ws_dd.cell(row=1, column=4).font = Font(name='Calibri', size=9, italic=True, color='888888')
ws_dd.merge_cells('A1:B1')

# Row 2: Subtitle
ws_dd.cell(row=2, column=1).value = '=TEXT(C1,"dd-mm-yyyy")'
ws_dd.cell(row=2, column=1).font = Font(name='Calibri', size=11, italic=True)
ws_dd.merge_cells('A2:G2')

# Row 3: Headers
headers_dd = ['Time Range', 'Project', 'Father Task', 'Sub Task Code', 'Category', 'Duration (h)', 'Description']
for i, h in enumerate(headers_dd, 1):
    ws_dd.cell(row=3, column=i, value=h)
style_header(ws_dd, 3, 7)

# Helper col H: MATCH to find first row with date = C1 (WPS-compatible)
# H4: first row, H5+: consecutive rows
dd_start = 4

# Hidden ref col I: total entry count for selected date
ws_dd.cell(row=dd_start, column=9).value = (
    f'=IFERROR(COUNTIF(\'Time Entries\'!A:A,$C$1),"")'
)
ws_dd.column_dimensions['I'].hidden = True

for r in range(dd_start, dd_start + DD_ROWS):
    row_offset = r - dd_start + 1
    if r == dd_start:
        # H4: MATCH finds first row with date = C1
        ws_dd.cell(row=r, column=8).value = (
            f'=IFERROR(MATCH($C$1,\'Time Entries\'!A:A,0),"")'
        )
    else:
        # H5+: consecutive rows (assumes entries for same date are consecutive)
        ws_dd.cell(row=r, column=8).value = (
            f'=IF(AND($I$4>=ROW()-{dd_start-1},$H${dd_start}<>""),$H${dd_start}+ROW()-{dd_start},"")'
        )

    # A: Time Range
    ws_dd.cell(row=r, column=1).value = (
        f'=IF($H{r}="","",IF(INDEX(\'Time Entries\'!G:G,$H{r})="","",'
        f'INDEX(\'Time Entries\'!G:G,$H{r})&" - "&INDEX(\'Time Entries\'!H:H,$H{r})))'
    )
    # B: Project
    ws_dd.cell(row=r, column=2).value = (
        f'=IF($H{r}="","",IF(INDEX(\'Time Entries\'!G:G,$H{r})="","",INDEX(\'Time Entries\'!M:M,$H{r})))'
    )
    # C: Father Task
    ws_dd.cell(row=r, column=3).value = (
        f'=IF($H{r}="","",IF(INDEX(\'Time Entries\'!G:G,$H{r})="","",INDEX(\'Time Entries\'!N:N,$H{r})))'
    )
    # D: Sub Task Code
    ws_dd.cell(row=r, column=4).value = (
        f'=IF($H{r}="","",IF(INDEX(\'Time Entries\'!G:G,$H{r})="","",INDEX(\'Time Entries\'!F:F,$H{r})))'
    )
    # E: Category
    ws_dd.cell(row=r, column=5).value = (
        f'=IF($H{r}="","",IF(INDEX(\'Time Entries\'!G:G,$H{r})="","",INDEX(\'Time Entries\'!G:G,$H{r})))'
    )
    # F: Duration
    ws_dd.cell(row=r, column=6).value = (
        f'=IF($H{r}="","",IF(INDEX(\'Time Entries\'!G:G,$H{r})="","",INDEX(\'Time Entries\'!J:J,$H{r})))'
    )
    ws_dd.cell(row=r, column=6).number_format = '0.00'
    # G: Description
    ws_dd.cell(row=r, column=7).value = (
        f'=IF($H{r}="","",IF(INDEX(\'Time Entries\'!G:G,$H{r})="","",INDEX(\'Time Entries\'!K:K,$H{r})))'
    )

    for c in range(1, 8):
        ws_dd.cell(row=r, column=c).font = NORMAL_FONT
        ws_dd.cell(row=r, column=c).border = THIN_BORDER
        ws_dd.cell(row=r, column=c).alignment = LEFT if c == 7 else CENTER

# Summary section
spacer = dd_start + DD_ROWS + 1

ws_dd.cell(row=spacer, column=1).value = 'DAY SUMMARY'
ws_dd.cell(row=spacer, column=1).font = Font(name='Calibri', size=12, bold=True)
ws_dd.cell(row=spacer, column=1).fill = LIGHT_BLUE_FILL
ws_dd.merge_cells(f'A{spacer}:G{spacer}')

r = spacer + 1
ws_dd.cell(row=r, column=1).value = 'Total Hours:'
ws_dd.cell(row=r, column=1).font = BOLD_FONT
ws_dd.cell(row=r, column=2).value = f'=SUM(F{dd_start}:F{dd_start+DD_ROWS-1})'
ws_dd.cell(row=r, column=2).number_format = '0.00'
ws_dd.cell(row=r, column=2).font = BOLD_FONT

r += 1
ws_dd.cell(row=r, column=1).value = 'Working Day:'
ws_dd.cell(row=r, column=1).font = BOLD_FONT
ws_dd.cell(row=r, column=2).value = '7.50 h'
ws_dd.cell(row=r, column=2).font = BOLD_FONT

r += 1
ws_dd.cell(row=r, column=1).value = '% Utilized:'
ws_dd.cell(row=r, column=1).font = BOLD_FONT
ws_dd.cell(row=r, column=2).value = f'=IF(F{spacer+1}=0,"N/A",F{spacer+1}/7.5)'
ws_dd.cell(row=r, column=2).number_format = '0.0%'
ws_dd.cell(row=r, column=2).font = BOLD_FONT

r += 1
ws_dd.cell(row=r, column=1).value = 'Status:'
ws_dd.cell(row=r, column=1).font = BOLD_FONT
formula_status = (
    f'=IF(F{spacer+1}=0,"No entries",'
    f'IF(F{spacer+1}>=7.5,"Full day",'
    f'"Partial: "&TEXT(7.5-F{spacer+1},"0.00")&"h remaining"))'
)
ws_dd.cell(row=r, column=2).value = formula_status
ws_dd.cell(row=r, column=2).font = BOLD_FONT

ws_dd.freeze_panes = 'A4'
print("Sheet 4 done: Daily Detail (WPS-compatible)")

# ===================================================================
# SHEET 5: Monthly Performance
# ===================================================================
ws_mo = wb.create_sheet('Monthly', 4)

ws_mo.column_dimensions['A'].width = 10
ws_mo.column_dimensions['B'].width = 35
ws_mo.column_dimensions['C'].width = 20
ws_mo.column_dimensions['D'].width = 14
ws_mo.column_dimensions['E'].width = 14
ws_mo.column_dimensions['F'].width = 16
ws_mo.column_dimensions['G'].width = 14
ws_mo.column_dimensions['H'].width = 16

# Hidden ref cells
ws_mo.cell(row=1, column=9).value = 'LatestDate'
ws_mo.cell(row=2, column=9).value = "=MAX('Time Entries'!A:A)"
ws_mo.cell(row=1, column=10).value = 'MonthStart'
ws_mo.cell(row=2, column=10).value = '=DATE(YEAR(I2),MONTH(I2),1)'
ws_mo.cell(row=1, column=11).value = 'MonthEnd'
ws_mo.cell(row=2, column=11).value = '=EOMONTH(I2,0)'
ws_mo.column_dimensions['I'].hidden = True
ws_mo.column_dimensions['J'].hidden = True
ws_mo.column_dimensions['K'].hidden = True

# Row 1: Title
ws_mo.cell(row=1, column=1).value = 'Monthly Performance'
ws_mo.cell(row=1, column=1).font = Font(name='Calibri', size=14, bold=True)
ws_mo.merge_cells('A1:H1')
ws_mo.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')

# Row 2: Month label
ws_mo.cell(row=2, column=1).value = '="Month: "&TEXT(J2,"mm-yyyy")'
ws_mo.cell(row=2, column=1).font = Font(name='Calibri', size=12, bold=True, italic=True)
ws_mo.merge_cells('A2:H2')

# Row 3: Headers
headers_mo = ['Code', 'Father Task', 'Project', 'Total Hours', 'Deadline (h)',
              '% Performance', 'Status', 'Completed Date']
for i, h in enumerate(headers_mo, 1):
    ws_mo.cell(row=3, column=i, value=h)
style_header(ws_mo, 3, 8)

mo_start = 4

# Pre-populated rows for each KPI
for r in range(mo_start, mo_start + MO_ROWS):
    row_offset = r - mo_start + 1

    # A: Code from KPIs (direct row reference, no AGGREGATE)
    kpi_src_row = r - mo_start + 2
    ws_mo.cell(row=r, column=1).value = (
        f'=IF(KPIs!A{kpi_src_row}="","",KPIs!A{kpi_src_row})'
    )

    # B: Father Task from KPIs
    ws_mo.cell(row=r, column=2).value = (
        f'=IF(KPIs!B{kpi_src_row}="","",KPIs!B{kpi_src_row})'
    )

    # C: Project
    ws_mo.cell(row=r, column=3).value = (
        f'=IF($B{r}="","",INDEX(KPIs!C:C,MATCH($B{r},KPIs!B:B,0)))'
    )

    # D: Total Hours in this month
    ws_mo.cell(row=r, column=4).value = (
        f'=IF($B{r}="","",SUMIFS(\'Time Entries\'!J:J,\'Time Entries\'!N:N,$B{r},'
        f"\'Time Entries\'!A:A,\">=\"&$J$2,\'Time Entries\'!A:A,\"<=\"&$K$2))"
    )
    ws_mo.cell(row=r, column=4).number_format = '0.00'

    # E: Deadline (h) from KPIs
    ws_mo.cell(row=r, column=5).value = (
        f'=IF($B{r}="","",INDEX(KPIs!F:F,MATCH($B{r},KPIs!B:B,0)))'
    )
    ws_mo.cell(row=r, column=5).number_format = '0.00'

    # F: % Performance (Use actual hours as-is — not a % of deadline)
    ws_mo.cell(row=r, column=6).value = (
        f'=IF(OR($B{r}="",E{r}=0),"",D{r}/E{r}*100)'
    )
    ws_mo.cell(row=r, column=6).number_format = '0.0"%"'

    # G: Status (from KPIs)
    ws_mo.cell(row=r, column=7).value = (
        f'=IF($B{r}="","",INDEX(KPIs!G:G,MATCH($B{r},KPIs!B:B,0)))'
    )

    # H: Completed Date (from KPIs)
    ws_mo.cell(row=r, column=8).value = (
        f'=IF($B{r}="","",INDEX(KPIs!H:H,MATCH($B{r},KPIs!B:B,0)))'
    )
    ws_mo.cell(row=r, column=8).number_format = 'dd-mm-yyyy'

    for c in range(1, 9):
        cell = ws_mo.cell(row=r, column=c)
        cell.font = NORMAL_FONT
        cell.alignment = CENTER if c >= 3 else LEFT
        cell.border = THIN_BORDER

# Summary section
mo_summary = mo_start + MO_ROWS + 1
ws_mo.cell(row=mo_summary, column=1).value = 'MONTHLY SUMMARY'
ws_mo.cell(row=mo_summary, column=1).font = Font(name='Calibri', size=12, bold=True)
ws_mo.cell(row=mo_summary, column=1).fill = LIGHT_BLUE_FILL
ws_mo.merge_cells(f'A{mo_summary}:H{mo_summary}')

r = mo_summary + 1
ws_mo.cell(row=r, column=1).value = 'Total Hours (all tasks):'
ws_mo.cell(row=r, column=1).font = BOLD_FONT
ws_mo.cell(row=r, column=2).value = f'=SUM(D{mo_start}:D{mo_start+MO_ROWS-1})'
ws_mo.cell(row=r, column=2).number_format = '0.00'

r += 1
ws_mo.cell(row=r, column=1).value = 'Avg Performance:'
ws_mo.cell(row=r, column=1).font = BOLD_FONT
ws_mo.cell(row=r, column=2).value = f'=AVERAGE(F{mo_start}:F{mo_start+MO_ROWS-1})'
ws_mo.cell(row=r, column=2).number_format = '0.0"%"'

r += 1
ws_mo.cell(row=r, column=1).value = 'Tasks On Time:'
ws_mo.cell(row=r, column=1).font = BOLD_FONT
ws_mo.cell(row=r, column=2).value = f'=COUNTIF(G{mo_start}:G{mo_start+MO_ROWS-1},"Done")'

r += 1
ws_mo.cell(row=r, column=1).value = 'Tasks Late:'
ws_mo.cell(row=r, column=1).font = BOLD_FONT
ws_mo.cell(row=r, column=2).value = f'=COUNTIF(G{mo_start}:G{mo_start+MO_ROWS-1},"Late")'

r += 1
ws_mo.cell(row=r, column=1).value = 'Overall:'
ws_mo.cell(row=r, column=1).font = BOLD_FONT
ws_mo.cell(row=r, column=2).value = (
    f'=IF(COUNTIF(G{mo_start}:G{mo_start+MO_ROWS-1},"Late")>0,'
    f'"Needs attention","All on track")'
)

ws_mo.freeze_panes = 'A4'
print("Sheet 5 done: Monthly (WPS-compatible)")

# ===================================================================
# Reorder sheets
# ===================================================================
desired_order = ['Time Entries', 'KPIs', 'Weekly Summary', 'Daily Detail', 'Monthly']
for i, name in enumerate(desired_order):
    idx = wb.sheetnames.index(name)
    wb.move_sheet(name, offset=i - idx)

output_path = 'worklog.xlsx'
wb.save(output_path)
print(f"\nWorkbook saved to {output_path}")
print("All done!")

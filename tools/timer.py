#!/usr/bin/env python3
"""
Work Log Timer CLI — for WPS Office users
Supports multiple independent timers running concurrently.
Usage:
  python timer.py start                  Start timer (always creates new)
  python timer.py stop [options]         Stop timer, write to worklog.xlsx
  python timer.py pomo [--work 25] [--rest 5]   Pomodoro session
  python timer.py today                  Show today's hours
  python timer.py week                   Show this week's hours
  python timer.py cancel                 Cancel running timer (discard)
  python timer.py status                 Show all timers
  python timer.py pause [-r REASON]      Pause active timer
  python timer.py continue               Resume paused timer
"""

import argparse
import json
import os
import sys
import time
import subprocess
import datetime
from pathlib import Path

import openpyxl
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))
import db
import webhook
import gsheets

STATE_FILE = Path.home() / '.worklog_timer.json'
WORKLOG_FILE = 'worklog.xlsx'

DAY_NAMES = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']


# ─── State ───────────────────────────────────────────────────────────

TIMER_ID_COUNTER = 0


def _next_timer_id(state=None):
    global TIMER_ID_COUNTER
    max_id = 0
    if state and 'timers' in state:
        for t in state['timers']:
            try:
                num = int(str(t.get('id', '')).lstrip('t#'))
                max_id = max(max_id, num)
            except Exception:
                pass
    TIMER_ID_COUNTER = max(TIMER_ID_COUNTER, max_id) + 1
    return f't{TIMER_ID_COUNTER}'


def load_state():
    if STATE_FILE.exists():
        data = json.loads(STATE_FILE.read_text(encoding='utf-8'))
        # Migrate old single-timer format
        if 'timers' not in data:
            old = dict(data)
            if 'segment_start' in old or old.get('accumulated_seconds', 0) > 0:
                old['id'] = 't0'
                old['pause_log'] = old.get('pause_log', [])
                old['pause_reason'] = old.get('pause_reason', '')
                return {'timers': [old]}
        return data
    return {'timers': []}


def save_state(data):
    STATE_FILE.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding='utf-8')


def clear_state():
    if STATE_FILE.exists():
        STATE_FILE.write_text(json.dumps({'timers': []}, indent=2), encoding='utf-8')


def get_elapsed(timer):
    """Get total elapsed seconds from a single timer dict."""
    acc = timer.get('accumulated_seconds', 0)
    seg = timer.get('segment_start')
    if seg and not timer.get('paused', False):
        sd = datetime.datetime.fromisoformat(seg)
        acc += (datetime.datetime.now() - sd).total_seconds()
    return acc


def parse_estimate(val, default=7.5):
    """Parse estimate string or float to float, defaulting to 7.5."""
    if val is None:
        return default
    try:
        res = float(val)
        return res if res > 0 else default
    except (ValueError, TypeError):
        return default


def sum_actual_hours(entries, task_name):
    """Sum actual logged duration for a given task/subtask name."""
    if not task_name:
        return 0.0
    t_name = str(task_name).strip().lower()
    return round(sum(
        float(e.get('duration', 0.0)) for e in entries if isinstance(e, dict)
        and (str(e.get('task', '') or e.get('subtask', '')).strip().lower() == t_name)
    ), 2)


def find_timer(state, project=None, task=None, subtask=None, paused=None, timer_id=None, user_id=None):
    """Find matching timer strictly by user_id. Returns most recent match or None."""
    timers = state.get('timers', [])
    candidates = list(timers)
    if user_id is not None:
        candidates = [t for t in candidates if str(t.get('user_id', '')) == str(user_id)]
        if not candidates:
            return None

    if timer_id is not None:
        tid_clean = str(timer_id).strip().lstrip('#').lower()
        candidates = [t for t in candidates if str(t.get('id', '')).strip().lstrip('#').lower() == tid_clean]
    if project is not None:
        p_clean = str(project).strip().lower()
        candidates = [t for t in candidates if str(t.get('project', '')).strip().lower() == p_clean]
    if task is not None:
        t_clean = str(task).strip().lower()
        candidates = [t for t in candidates if str(t.get('task', '')).strip().lower() == t_clean]
    if subtask is not None:
        s_clean = str(subtask).strip().lower()
        candidates = [
            t for t in candidates
            if str(t.get('subtask', '')).strip().lower() == s_clean or str(t.get('subtask_code', '')).strip().lower() == s_clean
        ]
    if paused is not None:
        candidates = [t for t in candidates if t.get('paused', False) == paused]
    return candidates[-1] if candidates else None



def find_timers(state, project=None, task=None, paused=None):
    """Find all matching timers."""
    timers = state.get('timers', [])
    candidates = list(timers)
    if project is not None:
        candidates = [t for t in candidates if t.get('project') == project]
    if task is not None:
        candidates = [t for t in candidates if t.get('task') == task]
    if paused is not None:
        candidates = [t for t in candidates if t.get('paused', False) == paused]
    return candidates


def remove_timer(state, timer_or_id):
    """Remove a timer by dict matching, object reference, or id. Returns the removed timer or None."""
    timers = state.get('timers', [])
    if isinstance(timer_or_id, dict):
        start_t = timer_or_id.get('segment_start')
        st_name = timer_or_id.get('subtask')
        for i, t in enumerate(timers):
            if t is timer_or_id or (start_t and t.get('segment_start') == start_t) or (st_name and t.get('subtask') == st_name):
                return timers.pop(i)
    for i, t in enumerate(timers):
        if str(t.get('id', '')) == str(timer_or_id):
            return timers.pop(i)
    return None


# ─── Git Auto-detect ─────────────────────────────────────────────────

def detect_project():
    try:
        top = subprocess.check_output(
            ['git', 'rev-parse', '--show-toplevel'],
            stderr=subprocess.DEVNULL
        ).decode().strip()
        return Path(top).name
    except Exception:
        return None


def detect_branch():
    try:
        return subprocess.check_output(
            ['git', 'branch', '--show-current'],
            stderr=subprocess.DEVNULL
        ).decode().strip()
    except Exception:
        return None


def detect_commit_msg():
    try:
        return subprocess.check_output(
            ['git', 'log', '-1', '--format=%s'],
            stderr=subprocess.DEVNULL
        ).decode().strip()
    except Exception:
        return None


def get_hours_for_subtask(subtask_name):
    """Total hours logged in Time Entries for a subtask."""
    if not subtask_name or not os.path.exists(WORKLOG_FILE):
        return 0.0
    wb = load_workbook(WORKLOG_FILE)
    ws = wb['Time Entries']
    total = 0.0
    s_clean = subtask_name.strip().lower()
    for r in range(2, ws.max_row + 1):
        st = ws.cell(r, 5).value
        if st and str(st).strip().lower() == s_clean:
            total += get_duration(ws, r)
    return total


def get_hours_for_kpi(kpi_name):
    """Total hours logged in Time Entries for a KPI / Father Task."""
    if not kpi_name or not os.path.exists(WORKLOG_FILE):
        return 0.0
    wb = load_workbook(WORKLOG_FILE)
    ws = wb['Time Entries']
    total = 0.0
    k_clean = kpi_name.strip().lower()
    for r in range(2, ws.max_row + 1):
        ft = ws.cell(r, 4).value
        if ft and str(ft).strip().lower() == k_clean:
            total += get_duration(ws, r)
    return total


def get_hours_for_project(project_name):
    """Total hours logged in Time Entries for a Project."""
    if not project_name or not os.path.exists(WORKLOG_FILE):
        return 0.0
    wb = load_workbook(WORKLOG_FILE)
    ws = wb['Time Entries']
    total = 0.0
    p_clean = project_name.strip().lower()
    for r in range(2, ws.max_row + 1):
        p = ws.cell(r, 3).value
        if p and str(p).strip().lower() == p_clean:
            total += get_duration(ws, r)
    return total


def get_hierarchy_watched_time(project, task, subtask, current_elapsed_sec=0):
    """Calculate watched time across all 3 levels: Subtask < KPI < Project."""
    cur_h = current_elapsed_sec / 3600.0
    sub_h = get_hours_for_subtask(subtask) + cur_h if subtask else 0.0
    kpi_h = get_hours_for_kpi(task) + cur_h if task else 0.0
    proj_h = get_hours_for_project(project) + cur_h if project else 0.0
    return {
        'subtask': (subtask, sub_h),
        'kpi': (task, kpi_h),
        'project': (project, proj_h),
    }


def cmd_performance(args=None):
    """Generate Friday Weekly Performance Report summarizing Subtask < KPI < Project time watching & progress."""
    if not os.path.exists(WORKLOG_FILE):
        print("No worklog.xlsx found.")
        return 1

    wb = load_workbook(WORKLOG_FILE)
    ws = wb['Time Entries']
    today = datetime.date.today()
    monday = today - datetime.timedelta(days=today.weekday())
    sunday = monday + datetime.timedelta(days=6)

    proj_hours = {}
    kpi_hours = {}
    sub_hours = {}
    total_week_hours = 0.0
    active_days = set()

    for r in range(2, ws.max_row + 1):
        d = ws.cell(r, 1).value
        if d is None:
            break
        if isinstance(d, datetime.datetime):
            d = d.date()
        if not (monday <= d <= sunday):
            continue
        active_days.add(d)
        dur = get_duration(ws, r)
        total_week_hours += dur

        p = str(ws.cell(r, 3).value or 'General')
        ft = str(ws.cell(r, 4).value or 'General')
        st = str(ws.cell(r, 5).value or '')

        proj_hours[p] = proj_hours.get(p, 0.0) + dur
        kpi_hours[ft] = kpi_hours.get(ft, 0.0) + dur
        if st:
            sub_hours[st] = sub_hours.get(st, 0.0) + dur

    subtasks = get_subtasks()
    completed_this_week = []
    in_progress_subtasks = []
    for s in subtasks:
        c_date = s.get('completed_date')
        if c_date:
            if hasattr(c_date, 'date'):
                c_date = c_date.date()
            if isinstance(c_date, datetime.date) and monday <= c_date <= sunday:
                completed_this_week.append(s)
        if str(s.get('status', '')).lower() in ('in progress', 'active'):
            in_progress_subtasks.append(s)

    target_hours = 37.5
    perf_pct = min(int((total_week_hours / target_hours) * 100), 100) if total_week_hours > 0 else 0

    lines = []
    lines.append(f"### 🏆 Friday Weekly Performance Report ({monday.strftime('%d-%m')} → {sunday.strftime('%d-%m-%Y')})")
    lines.append("")
    lines.append(f"**Weekly Watched Time:** `{total_week_hours:.2f}h / {target_hours:.1f}h` ({perf_pct}% of target)")
    lines.append(f"**Active Days:** `{len(active_days)}/5 days` │ **Daily Avg:** `{total_week_hours / max(len(active_days), 1):.2f}h/day`")
    lines.append(f"**Subtasks Completed This Week:** `{len(completed_this_week)}` │ **In Progress:** `{len(in_progress_subtasks)}`")
    lines.append("")

    lines.append("#### 📁 Level 1: Projects Watched Time")
    p_rows = []
    for p, h in sorted(proj_hours.items(), key=lambda x: -x[1]):
        p_rows.append([p, f"{h:.2f}h"])
    lines.append(_table(['Project', 'Watched Time'], p_rows, aligns=['<', '>']))
    lines.append("")

    lines.append("#### 🎯 Level 2: KPIs / Father Tasks Watched Time")
    k_rows = []
    for ft, h in sorted(kpi_hours.items(), key=lambda x: -x[1]):
        k_rows.append([ft, f"{h:.2f}h"])
    lines.append(_table(['Father Task', 'Watched Time'], k_rows, aligns=['<', '>']))
    lines.append("")

    lines.append("#### 🧩 Level 3: Subtasks Status & Progress")
    s_rows = []
    if completed_this_week:
        for s in completed_this_week:
            s_rows.append([f"✅ {s['code']}", s['name'], s['project'], 'Done'])
    for s in in_progress_subtasks:
        s_rows.append([f"🔵 {s['code']}", s['name'], s['project'], 'In Progress'])
    if s_rows:
        lines.append(_table(['Code', 'Subtask', 'Project', 'Status'], s_rows, aligns=['<', '<', '<', '<']))
    else:
        lines.append("- *(No subtask entries logged this week)*")

    report_text = "\n".join(lines)
    print(report_text)
    return 0


def _table(headers, rows, col_widths=None, aligns=None, title=None, footer=None, max_widths=None):
    """Render a GitHub-Flavored Markdown (GFM) pipe table for Discord and attached .md files."""
    n = len(headers)
    if aligns is None:
        aligns = ['<'] * n

    lines = []
    if title:
        lines.append(f"### {title}")
        lines.append("")

    lines.append("| " + " | ".join(headers) + " |")

    sep = []
    for a in aligns:
        if a == '>':
            sep.append("---:")
        elif a == '^':
            sep.append(":---:")
        else:
            sep.append(":---")
    lines.append("| " + " | ".join(sep) + " |")

    for row in rows:
        row_cells = []
        for i in range(n):
            val = row[i] if i < len(row) and row[i] is not None else ""
            row_cells.append(str(val))
        lines.append("| " + " | ".join(row_cells) + " |")

    if footer:
        footer_cells = []
        for i in range(n):
            val = footer[i] if i < len(footer) and footer[i] is not None else ""
            footer_cells.append(str(val))
        lines.append("| " + " | ".join(footer_cells) + " |")

    return "\n".join(lines)


def find_next_row(ws):
    """Find first empty row in Time Entries (col A check)."""
    for r in range(2, ws.max_row + 2):
        if ws.cell(r, 1).value is None:
            return r
    return ws.max_row + 1


def get_next_subtask_code(wb, father_task_name):
    """Generate next subtask code like PRJ1-KPI-1-ST-1 using KPI code from KPIs sheet."""
    kpi_code = get_kpi_code_from_task(father_task_name)
    if not kpi_code:
        return f'{father_task_name}-ST-1'
    ws = wb['Time Entries']
    max_st = 0
    for r2 in range(2, ws.max_row + 1):
        sc = ws.cell(r2, 6).value
        if sc and str(sc).startswith(f'{kpi_code}-ST-'):
            try:
                num = int(str(sc).rsplit('-', 1)[1])
                max_st = max(max_st, num)
            except:
                pass
    return f'{kpi_code}-ST-{max_st + 1}'


def append_entry(project, task, subtask, subtask_code, category, description,
                 start_dt, end_dt, break_info=None):
    """Append one row to Time Entries sheet and save."""
    wb = load_workbook(WORKLOG_FILE)
    ws = wb['Time Entries']
    r = find_next_row(ws)

    date_val = start_dt.date()
    day_str = DAY_NAMES[date_val.weekday()]

    ws.cell(r, 1).value = date_val
    ws.cell(r, 1).number_format = 'dd-mm-yyyy'
    ws.cell(r, 2).value = day_str
    ws.cell(r, 3).value = project
    ws.cell(r, 4).value = task
    ws.cell(r, 5).value = subtask or None
    ws.cell(r, 6).value = subtask_code or None
    ws.cell(r, 7).value = category or 'Work'
    ws.cell(r, 8).value = start_dt.time().replace(microsecond=0)
    ws.cell(r, 8).number_format = 'hh:mm'
    ws.cell(r, 9).value = end_dt.time().replace(microsecond=0)
    ws.cell(r, 9).number_format = 'hh:mm'
    ws.cell(r, 10).value = f'=IF(AND(H{r}<>"",I{r}<>""),(I{r}-H{r})*24,"")'
    ws.cell(r, 11).value = description or ''
    ws.cell(r, 12).value = break_info or ''
    # Helper cols M-R: let formulas auto-propagate (leave blank, formulas fill)

    db.insert_entry(project, task, subtask, subtask_code, category, description,
                    start_dt, end_dt, break_info or '')

    wb.save(WORKLOG_FILE)
    print(f"  Written to {WORKLOG_FILE} (row {r})")
    return r


# ─── Commands ────────────────────────────────────────────────────────

def cmd_start(args):
    subtask_name = getattr(args, 'subtask', None)
    subtask_code = getattr(args, 'subtask_code', None)
    user_id = getattr(args, 'user_id', None)
    user_name = getattr(args, 'user_name', None)

    # Validate mandatory parameters
    project_input = getattr(args, 'project', None) or detect_project()
    task_input = getattr(args, 'task', None) or detect_branch()
    category_input = getattr(args, 'category', None)
    estimate_input = getattr(args, 'estimate', None)

    if not project_input or not task_input or not subtask_name or not category_input or not estimate_input:
        print("Error: Missing mandatory parameters for start command. All of [project, task, subtask, category, estimate] are required.")
        return 1

    if subtask_name:
        subtasks = get_subtasks()
        match = None
        s_clean = subtask_name.strip().lower()
        for s in subtasks:
            if s['name'].strip().lower() == s_clean or s['code'].strip().lower() == s_clean:
                match = s
                break
        if not match:
            print(f"Error: Subtask '{subtask_name}' does not exist in SubTasks sheet.")
            print("Create it first using 'python -m tools.timer subtask add -s \"<name>\" -t \"<task>\"'")
            return 1
        subtask_name = match['name']
        subtask_code = match['code']
        if not getattr(args, 'task', None) and match.get('father_task'):
            args.task = match['father_task']
        if not getattr(args, 'project', None) and match.get('project'):
            args.project = match['project']

    state = load_state()
    target_project = args.project or detect_project()
    target_task = args.task or detect_branch()

    existing = find_timer(
        state,
        project=target_project,
        task=target_task,
        subtask=subtask_name,
        paused=False,
        user_id=str(user_id) if user_id else None
    )
    if existing:
        st_info = f"subtask '{subtask_name}'" if subtask_name else f"task '{target_task}'"
        u_info = f" for user {user_name}" if user_name else ""
        print(f"Error: A timer for {st_info} is already running (Timer #{existing['id']}){u_info}. Stop or pause it first.")
        return 1

    now = datetime.datetime.now()
    info = {
        'id': _next_timer_id(),
        'user_id': str(user_id) if user_id else None,
        'user_name': str(user_name) if user_name else None,
        'accumulated_seconds': 0,
        'segment_start': now.isoformat(),
        'paused': False,
        'project': target_project,
        'task': target_task,
        'subtask': subtask_name,
        'subtask_code': subtask_code,
        'category': getattr(args, 'category', None) or 'Development',
        'description': getattr(args, 'description', None) or detect_commit_msg(),
    }
    state.setdefault('timers', []).append(info)
    save_state(state)
    running = len(find_timers(state, paused=False))
    rows = []
    if info['user_name']:
        rows.append(['User', info['user_name']])
    rows.extend([
        ['Timer',    f"#{info['id']}"],
        ['Started',  now.strftime('%H:%M:%S')],
        ['Project',  info['project'] or '—'],
        ['Task',     info['task'] or '—'],
    ])
    if info['subtask']:
        sub_d = f"{info['subtask']} ({info['subtask_code']})" if info['subtask_code'] else info['subtask']
        rows.append(['Subtask', sub_d])
    rows.append(['Category', info['category'] or '—'])
    if info['description']:
        rows.append(['Desc', info['description']])
    rows.append(['Active', str(running)])
    print(_table(['Field', 'Value'], rows, [10, 38], title='▶ Timer Started'))
    if not info['project'] or not info['task']:
        print("(Use -p PROJECT -t TASK to override auto-detect)")
    return 0


def _safe_input(prompt, default=""):
    if not sys.stdin or not hasattr(sys.stdin, 'isatty') or not sys.stdin.isatty():
        return default
    try:
        val = input(prompt).strip()
        return val if val else default
    except (EOFError, KeyboardInterrupt, OSError):
        return default


def cmd_stop(args):
    state = load_state()
    timers = state.get('timers', [])
    if not timers:
        print("No timers running. Use 'start' first.")
        return 1

    stop_all = getattr(args, 'all', False) or getattr(args, 'stop_all', False)
    user_id = getattr(args, 'user_id', None)

    if stop_all:
        target_timers = list(timers)
        if user_id:
            user_t = [t for t in target_timers if str(t.get('user_id', '')) == str(user_id)]
            if user_t:
                target_timers = user_t

        count = 0
        for t in target_timers:
            dummy_args = argparse.Namespace(
                project=t.get('project'),
                task=t.get('task'),
                subtask=t.get('subtask'),
                subtask_code=t.get('subtask_code'),
                category=t.get('category'),
                description=t.get('description'),
                user_id=t.get('user_id'),
                user_name=t.get('user_name'),
                timer_id=t.get('id'),
                all=False,
                stop_all=False
            )
            ret = cmd_stop_single(dummy_args)
            if ret == 0:
                count += 1
        print(f"⏹️ Successfully stopped and saved {count} timer(s).")
        return 0

    return cmd_stop_single(args)


def cmd_stop_single(args):
    state = load_state()
    timers = state.get('timers', [])
    if not timers:
        print("No timers running. Use 'start' first.")
        return 1

    user_id = getattr(args, 'user_id', None)
    timer_id = getattr(args, 'timer_id', None)
    subtask = getattr(args, 'subtask', None)
    project = getattr(args, 'project', None)
    task = getattr(args, 'task', None)

    timer = None
    if timer_id:
        timer = find_timer(state, timer_id=timer_id)

    if not timer and subtask:
        if user_id:
            timer = find_timer(state, subtask=subtask, user_id=user_id)
        if not timer:
            timer = find_timer(state, subtask=subtask)

    if not timer and (project or task):
        if user_id:
            timer = find_timer(state, project=project, task=task, user_id=user_id)
        if not timer:
            timer = find_timer(state, project=project, task=task)

    if not timer and user_id:
        timer = find_timer(state, paused=False, user_id=user_id)
        if not timer:
            timer = find_timer(state, paused=True, user_id=user_id)
        if not timer:
            timer = find_timer(state, user_id=user_id)
        if not timer:
            print("No active or paused timers found for your account.")
            return 1

    if not timer and not user_id:
        timer = find_timer(state, paused=False)
        if not timer:
            timer = find_timer(state, paused=True)
        if not timer and timers:
            timer = timers[-1]

    if not timer:
        print("No timer found matching the criteria.")
        return 1

    now = datetime.datetime.now()
    acc = get_elapsed(timer)
    end_dt = now

    seg = timer.get('segment_start')
    if seg and not timer.get('paused', False):
        start_dt = datetime.datetime.fromisoformat(seg)
    else:
        start_dt = now - datetime.timedelta(seconds=acc)

    # Close any open pause
    pause_log = timer.get('pause_log', [])
    if timer.get('paused', False):
        pause_start = datetime.datetime.fromisoformat(timer['segment_start'])
        pause_duration = (now - pause_start).total_seconds() / 60
        reason = timer.get('pause_reason', 'break')
        pause_log.append({
            'reason': reason,
            'start': pause_start.isoformat(),
            'end': now.isoformat(),
            'duration_min': int(pause_duration)
        })

    # Build break info string
    break_info = ''
    total_break_min = 0
    if pause_log:
        parts = []
        for p in pause_log:
            parts.append(f"{p['reason']}({p['duration_min']}m)")
            total_break_min += p['duration_min']
        break_info = '; '.join(parts)

    project = args.project or timer.get('project') or detect_project() or _safe_input("Project: ", "General")
    task = args.task or timer.get('task') or detect_branch() or _safe_input("Father Task: ", "Task")
    subtask = args.subtask or timer.get('subtask')
    subtask_code = args.subtask_code or timer.get('subtask_code')
    category = args.category or timer.get('category') or 'Development'
    description = args.description or timer.get('description') or _safe_input("Description: ", "Work completed")
    u_name = timer.get('user_name') or getattr(args, 'user_name', None)
    if u_name and not description.startswith(f"[{u_name}]"):
        description = f"[{u_name}] {description}"

    if subtask and not subtask_code:
        wb_sc = load_workbook(WORKLOG_FILE)
        subtask_code = get_next_subtask_code(wb_sc, task)

    duration = acc / 3600
    tid = timer.get('id', '?')
    rows = []
    if u_name:
        rows.append(['User', u_name])
    rows.extend([
        ['Timer',    f'#{tid}'],
        ['Time',     f"{start_dt.strftime('%H:%M')} → {end_dt.strftime('%H:%M')}"],
        ['Duration', f'{duration:.2f}h'],
        ['Project',  project],
        ['Task',     task],
    ])
    if subtask:
        sub_d = f'{subtask} ({subtask_code})' if subtask_code else subtask
        rows.append(['Subtask', sub_d])
    rows.append(['Category', category])
    rows.append(['Desc', description])
    if break_info:
        rows.append(['Breaks', break_info])
        if total_break_min >= 60:
            rows.append(['Break total', f'{total_break_min}m ({total_break_min/60:.1f}h)'])
    print(_table(['Field', 'Value'], rows, [11, 38], title='■ Timer Stopped'))

    append_entry(project, task, subtask, subtask_code, category, description, start_dt, end_dt, break_info)
    gsheets.sync_time_entry(project, task, subtask, subtask_code, category, description,
                            start_dt.strftime('%H:%M'), end_dt.strftime('%H:%M'), duration,
                            break_info or '', start_dt.date())

    # Mark associated subtask as 'Hung' if not already Done
    if subtask:
        try:
            wb_sub = load_workbook(WORKLOG_FILE)
            ws_sub = wb_sub['SubTasks']
            for r_sub in range(2, ws_sub.max_row + 1):
                s_name = ws_sub.cell(r_sub, 2).value
                if s_name and str(s_name).strip() == subtask:
                    curr_st = str(ws_sub.cell(r_sub, 5).value or '')
                    if curr_st.lower() != 'done':
                        ws_sub.cell(r_sub, 5).value = 'Hung'
                        wb_sub.save(WORKLOG_FILE)
                        gsheets.sync_subtask_update(subtask, 'status', 'Hung')
                    break
        except Exception as e:
            print(f"Warning: Could not update subtask status to Hung: {e}")

    webhook.send(project=project, task=task, duration_h=duration,
                 description=description, category=category, subtask=subtask,
                 break_info=break_info,
                 start_time=start_dt.strftime('%H:%M'), end_time=end_dt.strftime('%H:%M'))
    remove_timer(state, timer)
    save_state(state)
    remaining = len(state.get('timers', []))
    if remaining:
        print(f"  {remaining} timer(s) still active")
    return 0


def cmd_cancel(args):
    state = load_state()
    timers = state.get('timers', [])
    if not timers:
        print("No timer running.")
        return 1
    user_id = getattr(args, 'user_id', None)
    timer = None
    if user_id:
        timer = find_timer(state, user_id=user_id, paused=False)
        if not timer:
            timer = find_timer(state, user_id=user_id)
        if not timer:
            print("No active or paused timers found for your account.")
            return 1
    else:
        timer = find_timer(state, paused=False)
        if not timer:
            timer = timers[-1]

    if not timer:
        print("No timer found to cancel.")
        return 1

    acc = get_elapsed(timer)
    tid = timer.get('id', '?')
    print(f"Cancelled timer #{tid} ({int(acc)//60}m discarded)")
    remove_timer(state, timer['id'])
    save_state(state)
    return 0


def cmd_status(args):
    state = load_state()
    timers = state.get('timers', [])
    user_id = getattr(args, 'user_id', None)
    if user_id:
        timers = [t for t in timers if str(t.get('user_id', '')) == str(user_id)]
    if not timers:
        print("No active or paused timers found for your account.")
        return 0

    today = datetime.date.today()
    timers_sorted = sorted(timers, key=lambda t: t.get('segment_start', ''))
    has_users = any(t.get('user_name') for t in timers)

    active_count = 0
    paused_count = 0
    hung_count = 0

    rows = []
    for t in timers_sorted:
        sd = datetime.datetime.fromisoformat(t['segment_start'])
        is_hung = sd.date() < today
        acc = get_elapsed(t)
        h, m = int(acc) // 3600, (int(acc) % 3600) // 60
        dur = f"{h}h{m:02d}m" if h else f"{m}m"

        proj = t.get('project', '?')
        task = t.get('task', '?')
        subtask = t.get('subtask', '') or '—'
        sub_code = t.get('subtask_code', '') or ''
        sub_str = f"{subtask}({sub_code})" if sub_code and subtask != '—' else subtask
        cat = t.get('category', '') or '—'
        desc = t.get('description', '') or '—'
        is_paused = t.get('paused', False)

        if is_hung and not is_paused:
            status = '⚠️ HUNG'
            hung_count += 1
        elif is_paused:
            reason = t.get('pause_reason', 'break')
            status = f'⏸️ {reason}'
            paused_count += 1
        else:
            status = '🟢 RUN'
            active_count += 1

        row = [t.get('id','?')]
        if has_users:
            row.append(t.get('user_name', '') or '—')
        row.extend([dur, sd.strftime('%H:%M'), status, proj, task, sub_str, cat, desc])
        rows.append(row)

    headers = ['ID', 'User', 'Time', 'Start', 'Status', 'Project', 'Task', 'Subtask', 'Category', 'Desc'] if has_users else ['ID', 'Time', 'Start', 'Status', 'Project', 'Task', 'Subtask', 'Category', 'Desc']
    widths = [3, 10, 6, 5, 9, 10, 18, 15, 10, 20] if has_users else [3, 6, 5, 9, 10, 18, 15, 10, 20]
    aligns = ['<'] * len(headers)

    print(_table(
        headers,
        rows,
        widths,
        aligns,
        title='Timer Status',
    ))
    print(f"\n🟢 Active: {active_count} │ ⏸️ Paused: {paused_count} │ ⚠️ Hung: {hung_count}")
    return 0


def cmd_pause(args):
    state = load_state()
    user_id = getattr(args, 'user_id', None)
    if user_id:
        timer = find_timer(state, project=getattr(args, 'project', None), task=getattr(args, 'task', None), subtask=getattr(args, 'subtask', None), user_id=user_id, paused=False)
        if not timer:
            timer = find_timer(state, user_id=user_id, paused=False)
        if not timer:
            print("No active timer to pause for your account.")
            return 1
    else:
        timer = find_timer(state, project=getattr(args, 'project', None), task=getattr(args, 'task', None), subtask=getattr(args, 'subtask', None), paused=False)
        if not timer:
            print("No active timer to pause.")
            return 1

    if timer.get('paused', False):
        print(f"Timer #{timer['id']} is already paused. Use 'continue' to resume.")
        return 1

    now = datetime.datetime.now()
    seg = datetime.datetime.fromisoformat(timer['segment_start'])
    elapsed = (now - seg).total_seconds()
    timer['accumulated_seconds'] = timer.get('accumulated_seconds', 0) + elapsed
    timer['segment_start'] = now.isoformat()
    timer['paused'] = True
    reason = args.reason if args.reason else 'break'
    timer['pause_reason'] = reason
    save_state(state)

    total = int(timer['accumulated_seconds']) // 60
    print(f"Timer #{timer['id']} paused at {now.strftime('%H:%M:%S')} ({total}m accumulated)")
    print(f"  Reason: {reason}")
    return 0


def cmd_continue(args):
    state = load_state()
    user_id = getattr(args, 'user_id', None)
    if user_id:
        timer = find_timer(state, project=getattr(args, 'project', None), task=getattr(args, 'task', None), subtask=getattr(args, 'subtask', None), user_id=user_id, paused=True)
        if not timer:
            timer = find_timer(state, user_id=user_id, paused=True)
        if not timer:
            print("No paused timer found for your account.")
            return 1
    else:
        timer = find_timer(state, project=getattr(args, 'project', None), task=getattr(args, 'task', None), subtask=getattr(args, 'subtask', None), paused=True)
        if not timer:
            print("No paused timer found.")
            return 1

    if not timer.get('paused', False):
        print(f"Timer #{timer['id']} is not paused.")
        return 1

    now = datetime.datetime.now()
    pause_start = datetime.datetime.fromisoformat(timer['segment_start'])
    pause_duration = (now - pause_start).total_seconds() / 60
    reason = timer.get('pause_reason', 'break')

    pause_log = timer.get('pause_log', [])
    pause_log.append({
        'reason': reason,
        'start': pause_start.isoformat(),
        'end': now.isoformat(),
        'duration_min': int(pause_duration)
    })
    timer['pause_log'] = pause_log
    timer['pause_reason'] = ''

    timer['segment_start'] = now.isoformat()
    timer['paused'] = False
    save_state(state)

    print(f"Timer #{timer['id']} resumed at {now.strftime('%H:%M:%S')} ({int(pause_duration)}m {reason})")
    if timer.get('project'):
        print(f"  Project: {timer['project']}")
    if timer.get('task'):
        print(f"  Task:    {timer['task']}")
    return 0


def parse_date_val(val):
    """Safely parse date value from cell (datetime, date, or str)."""
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        try:
            val_clean = val.split(' ')[0]
            if '-' in val_clean:
                parts = val_clean.split('-')
                if len(parts[0]) == 4:  # YYYY-MM-DD
                    return datetime.date(int(parts[0]), int(parts[1]), int(parts[2]))
                else:  # DD-MM-YYYY
                    return datetime.date(int(parts[2]), int(parts[1]), int(parts[0]))
        except Exception:
            pass
    return None


def get_duration(ws, r):
    """Get numeric duration from row. Computes from start/end if formula not evaluated."""
    dur = ws.cell(r, 10).value
    if isinstance(dur, (int, float)):
        return float(dur)
    start = ws.cell(r, 8).value
    end = ws.cell(r, 9).value
    if start and end:
        if isinstance(start, str):
            try:
                parts = start.strip().split(':')
                start = datetime.time(int(parts[0]), int(parts[1]), int(float(parts[2])) if len(parts) > 2 else 0)
            except Exception:
                pass
        if isinstance(end, str):
            try:
                parts = end.strip().split(':')
                end = datetime.time(int(parts[0]), int(parts[1]), int(float(parts[2])) if len(parts) > 2 else 0)
            except Exception:
                pass

        if isinstance(start, datetime.datetime):
            s = start
        elif isinstance(start, datetime.time):
            s = datetime.datetime.combine(datetime.date.today(), start)
        else:
            return 0.0

        if isinstance(end, datetime.datetime):
            e = end
        elif isinstance(end, datetime.time):
            e = datetime.datetime.combine(datetime.date.today(), end)
        else:
            return 0.0

        sec = (e - s).total_seconds()
        if sec < 0:
            sec += 86400  # Crosses midnight (e.g. 23:00 to 01:00)
        return sec / 3600.0
    return 0.0


def cmd_today(args):
    if not os.path.exists(WORKLOG_FILE):
        print(f"No {WORKLOG_FILE} found.")
        return 1
    wb = load_workbook(WORKLOG_FILE)
    ws = wb['Time Entries']
    today = datetime.date.today()
    total = 0.0

    entries = []
    for r in range(2, ws.max_row + 1):
        raw_d = ws.cell(r, 1).value
        d = parse_date_val(raw_d)
        if not d or d != today:
            continue
        dur = get_duration(ws, r)
        total += dur
        entries.append({
            'proj': str(ws.cell(r, 3).value or ''),
            'task': str(ws.cell(r, 4).value or ''),
            'desc': str(ws.cell(r, 11).value or ''),
            'dur': dur,
        })

    by_project = {}
    for e in entries:
        by_project.setdefault(e['proj'], []).append(e)

    for proj in sorted(by_project.keys()):
        proj_entries = by_project[proj]
        tbl_rows = []
        for e in proj_entries:
            tbl_rows.append([e['task'], f"{e['dur']:.2f}", e['desc']])
        proj_total = sum(e['dur'] for e in proj_entries)
        print(_table(
            ['Task', 'Hours', 'Description'],
            tbl_rows, [20, 6, 22],
            ['<', '>', '<'],
            title=f'Today {today.strftime("%d-%m-%Y")} [{proj}]',
            footer=['Subtotal', f'{proj_total:.2f}', ''],
        ))
        print()

    pct = int(min((total / 7.5) * 100, 100)) if total > 0 else 0
    filled = int(min((total / 7.5) * 14, 14)) if total > 0 else 0
    bar = ('█' * filled) + ('░' * (14 - filled))
    print(f"TOTAL: {total:.2f}h / 7.5h  {bar} {pct}%")
    return 0


def cmd_week(args):
    if not os.path.exists(WORKLOG_FILE):
        print(f"No {WORKLOG_FILE} found.")
        return 1
    wb = load_workbook(WORKLOG_FILE)
    ws = wb['Time Entries']

    today = datetime.date.today()
    monday = today - datetime.timedelta(days=today.weekday())
    sunday = monday + datetime.timedelta(days=6)

    by_project = {}
    total = 0.0
    days_set = set()
    for r in range(2, ws.max_row + 1):
        raw_d = ws.cell(r, 1).value
        d = parse_date_val(raw_d)
        if not d or not (monday <= d <= sunday):
            continue
        days_set.add(d)
        dur = get_duration(ws, r)
        total += dur
        proj = str(ws.cell(r, 3).value or '?')
        task = str(ws.cell(r, 4).value or '?')
        by_project.setdefault(proj, {})
        by_project[proj].setdefault(task, 0)
        by_project[proj][task] += dur

    for proj in sorted(by_project.keys()):
        tasks = by_project[proj]
        proj_total = sum(tasks.values())
        tbl_rows = []
        for t, h in sorted(tasks.items(), key=lambda x: -x[1]):
            tbl_rows.append([t, f'{h:.2f}'])
        print(_table(
            ['Task', 'Hours'],
            tbl_rows, [28, 7],
            ['<', '>'],
            title=f'Week {monday.strftime("%d-%m")} → {sunday.strftime("%d-%m-%Y")} [{proj}]',
            footer=['Subtotal', f'{proj_total:.2f}'],
        ))
        print()

    days_worked = len(days_set)
    avg = total / max(days_worked, 1)
    print(f"WEEK TOTAL: {total:.2f}h │ {days_worked} days │ Avg: {avg:.2f}h/day")
    return 0


def cmd_pomo(args):
    work_mins = args.work
    rest_mins = args.rest

    print(f"Pomodoro: {work_mins}min work + {rest_mins}min rest")
    print("  Press Ctrl+C to stop")

    cycles = 0
    while True:
        cycles += 1
        print(f"\n--- Cycle {cycles}: WORK {work_mins}min ---")
        try:
            countdown(work_mins * 60)
        except KeyboardInterrupt:
            print("\nPomodoro stopped.")
            break

        print(f"Work done! Rest {rest_mins}min...")
        try:
            countdown(rest_mins * 60, rest=True)
        except KeyboardInterrupt:
            print("\nPomodoro stopped.")
            break

    print(f"\nCompleted {cycles} pomodoro cycles.")
    return 0


# ─── Project helpers ─────────────────────────────────────────────────

def get_projects():
    """Read Projects sheet, return list of dicts."""
    if not os.path.exists(WORKLOG_FILE):
        return []
    wb = load_workbook(WORKLOG_FILE)
    ws = wb['Projects']
    projects = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 2).value
        if name is None or str(name).strip() in ('', '[Enter project name]'):
            break
        projects.append({
            'row': r,
            'code': str(ws.cell(r, 1).value or ''),
            'name': str(name).strip(),
            'description': str(ws.cell(r, 3).value or ''),
            'status': str(ws.cell(r, 4).value or 'Active'),
        })
    return projects


def get_project_code(project_name):
    """Get project code from Projects sheet by name."""
    projects = get_projects()
    for p in projects:
        if p['name'].lower() == project_name.lower():
            return p['code']
    return None


def get_kpi_code_from_task(task_name):
    """Get KPI code (e.g. PRJ1-KPI-1) from KPIs sheet by father task name."""
    if not os.path.exists(WORKLOG_FILE):
        return None
    wb = load_workbook(WORKLOG_FILE)
    ws = wb['KPIs']
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 2).value
        if name and str(name).strip() == task_name:
            return str(ws.cell(r, 1).value or '')
    return None


# ─── KPI commands ────────────────────────────────────────────────────

def get_kpis():
    """Read KPIs sheet, return list of dicts."""
    if not os.path.exists(WORKLOG_FILE):
        return []
    wb = load_workbook(WORKLOG_FILE)
    ws = wb['KPIs']
    kpis = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 2).value
        if name is None or name == '' or name == '[Enter father task name]':
            break
        kpis.append({
            'row': r,
            'code': str(ws.cell(r, 1).value or ''),
            'name': str(name),
            'project': str(ws.cell(r, 3).value or ''),
            'date': ws.cell(r, 4).value,
            'deadline_days': ws.cell(r, 5).value,
            'status': str(ws.cell(r, 7).value or ''),
            'completed': ws.cell(r, 8).value,
        })
    return kpis


def get_total_hours_for_task(task_name):
    """Get total logged hours for a father task from Time Entries."""
    if not os.path.exists(WORKLOG_FILE):
        return 0.0
    wb = load_workbook(WORKLOG_FILE)
    ws = wb['Time Entries']
    total = 0.0
    for r in range(2, ws.max_row + 1):
        t = ws.cell(r, 4).value
        if t is None:
            break
        if str(t) == task_name:
            total += get_duration(ws, r)
    return total


def cmd_kpi(args):
    if not args.kpi_cmd:
        print("Usage: kpi {list|add|edit|rename|delete|done|status}")
        print()
        print("  list       List all father tasks with status")
        print("  add        Add a new father task")
        print("  edit       Edit project/deadline")
        print("  rename     Rename father task (updates KPIs + Time Entries)")
        print("  delete     Delete a father task")
        print("  done       Mark a father task as completed")
        print("  status     Show overall task progress")
        return 1
    return args.kpi_func(args)


def cmd_kpi_list(args):
    kpis = get_kpis()
    if not kpis:
        print("No KPIs found. Add a project first with:")
        print("  python -m tools.timer project add -n \"My Project\"")
        print("Then add KPIs: kpi add -t NAME -p PROJECT -d DAYS")
        return 0

    by_project = {}
    for k in kpis:
        by_project.setdefault(k['project'], []).append(k)

    total_all = 0.0
    for proj_name in sorted(by_project.keys()):
        tbl_rows = []
        proj_total = 0.0
        for k in by_project[proj_name]:
            hours = get_total_hours_for_task(k['name'])
            proj_total += hours
            dl = k['deadline_days']
            dl_str = f"{dl}d" if dl else '—'
            st = k['status']
            if st.lower() == 'done':
                emoji = '✅'
            elif st.lower() in ('active', 'in progress'):
                emoji = '🔵'
            else:
                emoji = '⏳'
            tbl_rows.append([f'{emoji} {k["code"]}', k['name'], dl_str, f'{hours:.2f}', st])
        print(_table(
            ['Code', 'Task', 'Days', 'Hours', 'Status'],
            tbl_rows, [15, 20, 4, 6, 12],
            ['<', '<', '<', '>', '<'],
            title=f'KPIs [{proj_name}]',
            footer=['', 'Subtotal', '', f'{proj_total:.2f}', ''],
        ))
        print()
        total_all += proj_total
    print(f"Grand total: {total_all:.2f}h │ {len(kpis)} KPIs")
    return 0


def cmd_kpi_add(args):
    if not args.name:
        print("Error: -t NAME is required")
        return 1
    if not args.project:
        print("Error: -p PROJECT is required (add project first with 'project add')")
        return 1

    proj_code = get_project_code(args.project)
    if not proj_code:
        print(f"Error: Project '{args.project}' not found. Add it first with:")
        print(f"  python -m tools.timer project add -n \"{args.project}\"")
        return 1

    wb = load_workbook(WORKLOG_FILE)
    ws = wb['KPIs']
    r = 2
    while ws.cell(r, 2).value and str(ws.cell(r, 2).value).strip() not in ('', '[Enter father task name]'):
        if str(ws.cell(r, 2).value).strip() == args.name.strip():
            print(f"KPI '{args.name}' already exists at row {r}.")
            return 1
        r += 1

    # Count existing KPIs for this project to get next KPI number
    max_kpi_num = 0
    for r_check in range(2, ws.max_row + 1):
        code_val = ws.cell(r_check, 1).value
        if code_val and str(code_val).startswith(f'{proj_code}-KPI-'):
            try:
                num = int(str(code_val).rsplit('-', 1)[1])
                max_kpi_num = max(max_kpi_num, num)
            except:
                pass
    new_code = f'{proj_code}-KPI-{max_kpi_num + 1}'

    ws.cell(r, 1).value = new_code
    ws.cell(r, 2).value = args.name
    ws.cell(r, 3).value = args.project or ''
    ws.cell(r, 4).value = datetime.date.today()
    ws.cell(r, 4).number_format = 'dd-mm-yyyy'
    ws.cell(r, 5).value = args.deadline
    ws.cell(r, 6).value = f'=IF(E{r}<>"",E{r}*7.5,"")'
    ws.cell(r, 7).value = 'In Progress'
    ws.cell(r, 8).value = None
    wb.save(WORKLOG_FILE)
    print(f"KPI added: {args.name} (code={new_code}, project={args.project}, deadline={args.deadline}d)")
    gsheets.sync_kpi_add(new_code, args.name, args.project, args.deadline)
    return 0


def cmd_kpi_done(args):
    if not args.name:
        print("Error: -t NAME is required")
        return 1
    kpis = get_kpis()
    found = [k for k in kpis if k['name'] == args.name]
    if not found:
        print(f"KPI '{args.name}' not found.")
        return 1

    # Stop any active or paused timer running for this KPI/Task first
    state = load_state()
    timer = find_timer(state, task=args.name)
    if timer:
        dummy_stop_args = argparse.Namespace(
            timer_id=timer.get('id'),
            subtask=timer.get('subtask'),
            project=timer.get('project'),
            task=args.name,
            user_id=timer.get('user_id'),
            user_name=timer.get('user_name'),
            category=timer.get('category'),
            description=timer.get('description'),
            all=False,
            stop_all=False
        )
        cmd_stop_single(dummy_stop_args)

    wb = load_workbook(WORKLOG_FILE)
    ws = wb['KPIs']
    r = found[0]['row']
    ws.cell(r, 7).value = 'Done'
    ws.cell(r, 8).value = datetime.date.today()
    ws.cell(r, 8).number_format = 'dd-mm-yyyy'
    wb.save(WORKLOG_FILE)
    print(f"KPI '{args.name}' marked as Done.")
    gsheets.sync_kpi_update(args.name, 'status', 'Done')
    gsheets.sync_kpi_update(args.name, 'completed_date', datetime.date.today().strftime('%Y-%m-%d'))
    return 0


def cmd_kpi_edit(args):
    kpis = get_kpis()
    found = [k for k in kpis if k['name'] == args.name]
    if not found:
        print(f"KPI '{args.name}' not found.")
        return 1
    wb = load_workbook(WORKLOG_FILE)
    ws = wb['KPIs']
    r = found[0]['row']
    changed = []
    if args.project is not None:
        ws.cell(r, 3).value = args.project
        changed.append(f"project={args.project}")
    if args.deadline is not None:
        ws.cell(r, 5).value = args.deadline
        changed.append(f"deadline={args.deadline}d")
    wb.save(WORKLOG_FILE)
    if changed:
        print(f"KPI '{args.name}' updated: {', '.join(changed)}")
        if args.project:
            gsheets.sync_kpi_update(args.name, 'project', args.project)
        if args.deadline:
            gsheets.sync_kpi_update(args.name, 'deadline_days', str(args.deadline))
    else:
        print("No changes made. Use -p PROJECT and/or -d DEADLINE.")
    return 0


def cmd_kpi_rename(args):
    kpis = get_kpis()
    found = [k for k in kpis if k['name'] == args.name]
    if not found:
        print(f"KPI '{args.name}' not found.")
        return 1
    new_name = args.new_name or args.name
    if new_name == args.name:
        print("New name is same as old name.")
        return 0
    wb = load_workbook(WORKLOG_FILE)
    ws_kpi = wb['KPIs']
    ws_st = wb['SubTasks']
    ws_te = wb['Time Entries']
    r = found[0]['row']
    ws_kpi.cell(r, 2).value = new_name

    for r2 in range(2, ws_st.max_row + 1):
        old_val = ws_st.cell(r2, 3).value
        if old_val and str(old_val).strip().lower() == args.name.strip().lower():
            ws_st.cell(r2, 3).value = new_name

    for r2 in range(2, ws_te.max_row + 1):
        old_val = ws_te.cell(r2, 4).value
        if old_val and str(old_val).strip().lower() == args.name.strip().lower():
            ws_te.cell(r2, 4).value = new_name

    wb.save(WORKLOG_FILE)
    print(f"Renamed KPI '{args.name}' -> '{new_name}' (updated in KPIs, SubTasks, Time Entries)")
    gsheets.sync_kpi_rename(args.name, new_name)
    gsheets.sync_kpi_rename_propagate(args.name, new_name)
    return 0


def cmd_kpi_delete(args):
    kpis = get_kpis()
    found = [k for k in kpis if k['name'] == args.name]
    if not found:
        print(f"KPI '{args.name}' not found.")
        return 1
    wb = load_workbook(WORKLOG_FILE)
    ws = wb['KPIs']
    r = found[0]['row']
    ws.delete_rows(r)
    wb.save(WORKLOG_FILE)
    print(f"KPI '{args.name}' deleted.")
    gsheets.sync_kpi_delete(args.name)
    return 0


def cmd_kpi_status(args):
    kpis = get_kpis()
    if not kpis:
        print("No KPIs found.")
        return 0

    by_project = {}
    for k in kpis:
        by_project.setdefault(k['project'], []).append(k)

    total_all = 0.0
    done_all = 0
    for proj_name in sorted(by_project.keys()):
        proj_kpis = by_project[proj_name]
        print(f"\n[{proj_name}]")
        print(f"{'Code':<18} {'Task':<25} {'Hours':>8} {'Deadline':<10} {'Status':<12} {'Completed':<15}")
        print("-" * 90)
        proj_total = 0.0
        for k in proj_kpis:
            hours = get_total_hours_for_task(k['name'])
            proj_total += hours
            dl = k['deadline_days']
            dl_str = f"{dl}d" if dl else '—'
            comp = ''
            if k['completed']:
                comp = k['completed'].strftime('%d-%m-%Y') if hasattr(k['completed'], 'strftime') else str(k['completed'])
            print(f"{k['code']:<18} {k['name']:<25} {hours:>8.2f} {dl_str:<10} {k['status']:<12} {comp:<15}")
        done_proj = sum(1 for k in proj_kpis if k['status'] == 'Done')
        print(f"{'Project total':<18} {'':25} {proj_total:>8.2f} {'':10} {'Done:' + str(done_proj) + '/' + str(len(proj_kpis)):<12}")
        total_all += proj_total
        done_all += done_proj

    print(f"\nOverall: {done_all}/{len(kpis)} done | Total: {total_all:.2f}h")
    return 0


# ─── Tasks command ───────────────────────────────────────────────────

def cmd_tasks(args):
    if not os.path.exists(WORKLOG_FILE):
        print(f"No {WORKLOG_FILE} found.")
        return 1
    wb = load_workbook(WORKLOG_FILE)
    ws = wb['Time Entries']

    groups = {}
    for r in range(2, ws.max_row + 1):
        ft = ws.cell(r, 4).value
        if ft is None or str(ft).strip() == '':
            break
        ft = str(ft).strip()
        proj = str(ws.cell(r, 3).value or '')
        st = ws.cell(r, 5).value
        sc = ws.cell(r, 6).value
        cat = ws.cell(r, 7).value or ''
        desc = ws.cell(r, 11).value or ''
        dur = get_duration(ws, r)
        groups.setdefault(ft, []).append({'proj': proj, 'sub': str(st or ''), 'sc': str(sc or ''), 'cat': str(cat), 'dur': dur, 'desc': str(desc)})

    if args.task:
        ft = args.task
        entries = groups.get(ft, [])
        if not entries:
            print(f"No entries for father task '{ft}'.")
            return 0
        total = sum(e['dur'] for e in entries)
        proj = entries[0]['proj'] if entries else ''
        tbl_rows = []
        for e in entries:
            tbl_rows.append([e['sub'] or '—', e['sc'] or '—', e['cat'], f"{e['dur']:.2f}", e['desc']])
        print(_table(
            ['SubTask', 'Code', 'Category', 'Hours', 'Description'],
            tbl_rows, [16, 13, 12, 6, 22],
            ['<', '<', '<', '>', '<'],
            title=f'{ft} [{proj}]',
            footer=['TOTAL', '', '', f'{total:.2f}', ''],
        ))
    else:
        if not groups:
            print("No time entries found.")
            return 0

        # Group by project
        by_project = {}
        for ft, entries in groups.items():
            proj = entries[0]['proj'] if entries else '?'
            by_project.setdefault(proj, {})[ft] = entries

        total_all = 0.0
        for proj in sorted(by_project.keys()):
            tasks = by_project[proj]
            tbl_rows = []
            proj_total = 0.0
            for ft, entries in sorted(tasks.items()):
                total = sum(e['dur'] for e in entries)
                proj_total += total
                tbl_rows.append([ft, str(len(entries)), f'{total:.2f}'])
            print(_table(
                ['Father Task', 'Entries', 'Hours'],
                tbl_rows, [25, 7, 7],
                ['<', '>', '>'],
                title=f'Tasks [{proj}]',
                footer=['Total', '', f'{proj_total:.2f}'],
            ))
            print()
            total_all += proj_total
        print(f"Grand total: {total_all:.2f}h")
    return 0


# ─── Subtask helpers ─────────────────────────────────────────────────

def get_subtasks():
    """Read SubTasks sheet, return list of dicts."""
    if not os.path.exists(WORKLOG_FILE):
        return []
    wb = load_workbook(WORKLOG_FILE)
    ws = wb['SubTasks']
    subtasks = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 2).value
        if name is None or name == '' or name == '[Enter subtask name]':
            break
        subtasks.append({
            'row': r,
            'code': str(ws.cell(r, 1).value or ''),
            'name': str(name),
            'father_task': str(ws.cell(r, 3).value or ''),
            'project': str(ws.cell(r, 4).value or ''),
            'status': str(ws.cell(r, 5).value or ''),
            'created_date': ws.cell(r, 6).value,
            'completed_date': ws.cell(r, 7).value,
            'notes': str(ws.cell(r, 8).value or ''),
        })
    return subtasks


def get_next_subtask_code_from_sheet(father_task_name, project):
    """Get next subtask code like PRJ-1-KPI-1-ST-1 from SubTasks sheet."""
    kpi_code = get_kpi_code_from_task(father_task_name)
    if not kpi_code:
        return None
    max_n = 0
    subtasks = get_subtasks()
    for s in subtasks:
        if s['code'].startswith(f'{kpi_code}-ST-'):
            try:
                n = int(s['code'].rsplit('-', 1)[1])
                max_n = max(max_n, n)
            except:
                pass
    return f'{kpi_code}-ST-{max_n + 1}'


# ─── Subtask ─────────────────────────────────────────────────────────

def cmd_subtask(args):
    if not args.sub_cmd:
        print("subtask: add | list | edit | rename | delete | done | status | start")
        print()
        print("  add        Add a new subtask to SubTasks sheet")
        print("  list       List all subtasks")
        print("  edit       Edit subtask project/father task")
        print("  rename     Rename subtask (updates SubTasks + Time Entries)")
        print("  delete     Delete subtask from SubTasks sheet")
        print("  done       Mark a subtask as completed")
        print("  status     Show subtask progress")
        print("  start      Start timer for an existing subtask")
        return 1
    return args.sub_func(args)


def cmd_sub_start(args):
    if not getattr(args, 'subtask_code', None):
        args.subtask_code = None
    return cmd_start(args)


def cmd_sub_list(args):
    if not os.path.exists(WORKLOG_FILE):
        print(f"No {WORKLOG_FILE} found.")
        return 1
        
    subtasks = get_subtasks()
    st_status = {}
    if subtasks:
        for s in subtasks:
            st_status[s['name']] = s['status']

    wb = load_workbook(WORKLOG_FILE)
    ws = wb['Time Entries']
    seen = {}
    for r in range(2, ws.max_row + 1):
        ft = ws.cell(r, 4).value
        st = ws.cell(r, 5).value
        if not ft or not st:
            continue
        ft = str(ft).strip()
        st = str(st).strip()
        if args.task and args.task.lower() not in ft.lower():
            continue
        dur = get_duration(ws, r)
        seen.setdefault(st, {'father_tasks': set(), 'total_hours': 0.0})
        seen[st]['father_tasks'].add(ft)
        seen[st]['total_hours'] += dur
        
    if not seen:
        print("No subtasks found.")
        return 0

    tbl_rows = []
    for name, info in sorted(seen.items()):
        fts = ', '.join(sorted(info['father_tasks']))
        if len(fts) > 20:
            fts = fts[:17] + '...'
        status = st_status.get(name, '')
        if status:
            if status.lower() == 'done':
                emoji = '✅'
            elif status.lower() in ('active', 'in progress'):
                emoji = '🔵'
            else:
                emoji = '⏳'
        else:
            emoji = '⚪'
        tbl_rows.append([f'{emoji} {name}', fts, status or '—', f"{info['total_hours']:.2f}"])
    print(_table(
        ['Subtask', 'Father Task', 'Status', 'Hours'],
        tbl_rows, [22, 20, 11, 6],
        ['<', '<', '<', '>'],
        title='Subtasks',
    ))
    return 0


def cmd_sub_rename(args):
    if not os.path.exists(WORKLOG_FILE):
        print(f"No {WORKLOG_FILE} found.")
        return 1
    wb = load_workbook(WORKLOG_FILE)
    ws_st = wb['SubTasks']
    ws_te = wb['Time Entries']

    # Update SubTasks sheet definition
    s_clean = args.subtask.strip().lower()
    st_count = 0
    for r in range(2, ws_st.max_row + 1):
        s_name = ws_st.cell(r, 2).value
        s_code = ws_st.cell(r, 1).value
        if s_name and (str(s_name).strip().lower() == s_clean or str(s_code or '').strip().lower() == s_clean):
            ws_st.cell(r, 2).value = args.new_name
            st_count += 1

    # Update Time Entries sheet
    te_count = 0
    for r in range(2, ws_te.max_row + 1):
        st = ws_te.cell(r, 5).value
        sc = ws_te.cell(r, 6).value
        if st and (str(st).strip().lower() == s_clean or str(sc or '').strip().lower() == s_clean):
            ws_te.cell(r, 5).value = args.new_name
            te_count += 1

    wb.save(WORKLOG_FILE)
    print(f"Renamed subtask '{args.subtask}' -> '{args.new_name}' (updated in SubTasks and {te_count} Time Entries).")
    gsheets.sync_subtask_rename(args.subtask, args.new_name)
    return 0


def cmd_sub_delete(args):
    if not os.path.exists(WORKLOG_FILE):
        print(f"No {WORKLOG_FILE} found.")
        return 1
    wb = load_workbook(WORKLOG_FILE)
    ws_st = wb['SubTasks']
    ws_te = wb['Time Entries']

    s_clean = args.subtask.strip().lower()
    # Delete row in SubTasks sheet
    rows_to_delete = []
    for r in range(2, ws_st.max_row + 1):
        s_name = ws_st.cell(r, 2).value
        s_code = ws_st.cell(r, 1).value
        if s_name and (str(s_name).strip().lower() == s_clean or str(s_code or '').strip().lower() == s_clean):
            rows_to_delete.append(r)

    for r in reversed(rows_to_delete):
        ws_st.delete_rows(r)

    # Clear matching Time Entries subtask field
    te_count = 0
    for r in range(2, ws_te.max_row + 1):
        st = ws_te.cell(r, 5).value
        sc = ws_te.cell(r, 6).value
        if st and (str(st).strip().lower() == s_clean or str(sc or '').strip().lower() == s_clean):
            ws_te.cell(r, 5).value = None
            te_count += 1

    wb.save(WORKLOG_FILE)
    print(f"Deleted subtask '{args.subtask}' ({len(rows_to_delete)} subtask row removed, {te_count} time entries unlinked).")
    gsheets.sync_subtask_delete(args.subtask)
    return 0


def cmd_sub_add(args):
    if not args.subtask:
        print("Error: -s NAME is required")
        return 1
    if not args.task:
        print("Error: -t FATHER_TASK is required")
        return 1
    if not getattr(args, 'project', None):
        print("Error: -p PROJECT is required")
        return 1
    if not getattr(args, 'estimate', None):
        print("Error: -e ESTIMATE is required")
        return 1

    # Check if already exists
    subtasks = get_subtasks()
    for s in subtasks:
        if s['name'].lower() == args.subtask.strip().lower():
            print(f"Subtask '{args.subtask}' already exists (row {s['row']}).")
            return 1

    wb = load_workbook(WORKLOG_FILE)
    ws = wb['SubTasks']
    r = 2
    while ws.cell(r, 2).value and str(ws.cell(r, 2).value).strip() not in ('', '[Enter subtask name]'):
        r += 1

    # Detect project from args or father task lookup
    project = args.project
    if not project:
        kpis = get_kpis()
        for k in kpis:
            if k['name'] == args.task:
                project = k['project']
                break
    if not project:
        print(f"Error: Could not determine project for father task '{args.task}'. Use -p PROJECT.")
        return 1

    code = get_next_subtask_code_from_sheet(args.task, project)
    if not code:
        print(f"Error: Could not generate subtask code. Is father task '{args.task}' valid?")
        return 1

    ws.cell(r, 1).value = code
    ws.cell(r, 2).value = args.subtask.strip()
    ws.cell(r, 3).value = args.task
    ws.cell(r, 4).value = project
    ws.cell(r, 5).value = 'In Progress'
    ws.cell(r, 6).value = datetime.date.today()
    ws.cell(r, 6).number_format = 'dd-mm-yyyy'
    ws.cell(r, 7).value = None
    ws.cell(r, 8).value = args.notes or ''
    wb.save(WORKLOG_FILE)
    print(f"Subtask added: {args.subtask} (code={code}, father={args.task}, project={project})")
    gsheets.sync_subtask_add(code, args.subtask, args.task, project)
    return 0


def cmd_sub_edit(args):
    subtasks = get_subtasks()
    found = [s for s in subtasks if s['name'] == args.subtask]
    if not found:
        print(f"Subtask '{args.subtask}' not found.")
        return 1
    wb = load_workbook(WORKLOG_FILE)
    ws = wb['SubTasks']
    r = found[0]['row']
    changed = []
    if args.new_name is not None:
        ws.cell(r, 2).value = args.new_name
        changed.append(f"name={args.new_name}")
        ws_te = wb['Time Entries']
        s_clean = args.subtask.strip().lower()
        for r2 in range(2, ws_te.max_row + 1):
            st = ws_te.cell(r2, 5).value
            if st and str(st).strip().lower() == s_clean:
                ws_te.cell(r2, 5).value = args.new_name
    if args.task is not None:
        ws.cell(r, 3).value = args.task
        changed.append(f"father_task={args.task}")
    if args.project is not None:
        ws.cell(r, 4).value = args.project
        changed.append(f"project={args.project}")
    if getattr(args, 'estimate', None) is not None:
        est_val = parse_estimate(args.estimate)
        ws.cell(r, 9).value = est_val
        changed.append(f"estimate={est_val:.1f}h")
    if args.notes is not None:
        ws.cell(r, 8).value = args.notes
        changed.append("notes updated")
    wb.save(WORKLOG_FILE)
    if changed:
        print(f"Subtask '{args.subtask}' updated: {', '.join(changed)}")
        if args.task:
            gsheets.sync_subtask_update(args.subtask, 'father_task', args.task)
        if args.project:
            gsheets.sync_subtask_update(args.subtask, 'project', args.project)
    else:
        print("No changes made. Use -t FATHER_TASK, -p PROJECT, -n NEW_NAME, or --notes.")
    return 0


def cmd_sub_done(args):
    subtasks = get_subtasks()
    found = [s for s in subtasks if s['name'] == args.subtask]
    if not found:
        print(f"Subtask '{args.subtask}' not found.")
        return 1

    # Stop any active or paused timer running for this subtask first
    state = load_state()
    timer = find_timer(state, subtask=args.subtask)
    if timer:
        dummy_stop_args = argparse.Namespace(
            timer_id=timer.get('id'),
            subtask=args.subtask,
            project=timer.get('project'),
            task=timer.get('task'),
            user_id=timer.get('user_id'),
            user_name=timer.get('user_name'),
            category=timer.get('category'),
            description=timer.get('description'),
            all=False,
            stop_all=False
        )
        cmd_stop_single(dummy_stop_args)

    wb = load_workbook(WORKLOG_FILE)
    ws = wb['SubTasks']
    r = found[0]['row']
    ws.cell(r, 5).value = 'Done'
    ws.cell(r, 7).value = datetime.date.today()
    ws.cell(r, 7).number_format = 'dd-mm-yyyy'
    wb.save(WORKLOG_FILE)
    print(f"Subtask '{args.subtask}' marked as Done.")
    gsheets.sync_subtask_update(args.subtask, 'status', 'Done')
    gsheets.sync_subtask_update(args.subtask, 'completed_date', datetime.date.today().strftime('%Y-%m-%d'))
    return 0


def cmd_sub_status(args):
    subtasks = get_subtasks()
    if not subtasks:
        print("No subtasks found. Add one with: subtask add -s NAME -t FATHER_TASK")
        return 0

    by_father = {}
    for s in subtasks:
        key = f"{s['project']} / {s['father_task']}"
        by_father.setdefault(key, []).append(s)

    total_all = 0
    done_all = 0
    for key in sorted(by_father.keys()):
        items = by_father[key]
        tbl_rows = []
        for s in items:
            comp = ''
            if s['completed_date']:
                comp = s['completed_date'].strftime('%d-%m-%Y') if hasattr(s['completed_date'], 'strftime') else str(s['completed_date'])
            st = s['status']
            if st.lower() == 'done':
                emoji = '✅'
            elif st.lower() in ('active', 'in progress'):
                emoji = '🔵'
            else:
                emoji = '⏳'
            tbl_rows.append([f'{emoji} {s["code"]}', s['name'], st, comp or '—'])
        done_sub = sum(1 for s in items if s['status'] == 'Done')
        print(_table(
            ['Code', 'Subtask', 'Status', 'Completed'],
            tbl_rows, [15, 20, 11, 11],
            title=f'Subtask Status [{key}]',
            footer=['', f'Done: {done_sub}/{len(items)}', '', ''],
        ))
        print()
        total_all += len(items)
        done_all += done_sub

    print(f"Overall: {done_all}/{total_all} subtasks done")
    return 0


# ─── Project commands ────────────────────────────────────────────────

def cmd_project(args):
    if not args.proj_cmd:
        print("Usage: project {list|add|rename|archive}")
        print()
        print("  list         List all projects")
        print("  add          Add a new project")
        print("  rename       Rename a project (updates Projects + KPIs + Time Entries)")
        print("  archive      Archive a project")
        return 1
    return args.proj_func(args)


def get_next_project_code():
    """Generate next project code PRJ-1, PRJ-2..."""
    projects = get_projects()
    max_n = 0
    for p in projects:
        code = p['code']
        if code and code.startswith('PRJ-'):
            try:
                n = int(code.split('-')[1])
                max_n = max(max_n, n)
            except:
                pass
    return f'PRJ-{max_n + 1}'


def cmd_project_add(args):
    if not args.name:
        print("Error: -n NAME is required")
        return 1
    wb = load_workbook(WORKLOG_FILE)
    ws = wb['Projects']
    r = 2
    while ws.cell(r, 2).value and str(ws.cell(r, 2).value).strip() not in ('', '[Enter project name]'):
        if str(ws.cell(r, 2).value).strip().lower() == args.name.strip().lower():
            print(f"Project '{args.name}' already exists at row {r}.")
            return 1
        r += 1

    code = get_next_project_code()
    ws.cell(r, 1).value = code
    ws.cell(r, 2).value = args.name.strip()
    ws.cell(r, 3).value = args.description or ''
    ws.cell(r, 4).value = 'Active'
    ws.cell(r, 5).value = datetime.date.today()
    ws.cell(r, 5).number_format = 'dd-mm-yyyy'
    wb.save(WORKLOG_FILE)
    print(f"Project added: {args.name} (code={code})")
    gsheets.sync_project_add(code, args.name, args.description)
    return 0


def cmd_project_list(args):
    projects = get_projects()
    if not projects:
        print("No projects found. Add one with: project add -n NAME")
        return 0
    print(f"{'Code':<10} {'Name':<25} {'Description':<40} {'Status':<12} {'KPIs':>6}")
    print("-" * 95)
    kpis = get_kpis()
    for p in projects:
        kpi_count = sum(1 for k in kpis if k['project'].lower() == p['name'].lower())
        desc = p['description'][:38] + '..' if len(p['description']) > 38 else p['description']
        print(f"{p['code']:<10} {p['name']:<25} {desc:<40} {p['status']:<12} {kpi_count:>6}")
    return 0


def cmd_project_rename(args):
    projects = get_projects()
    found = [p for p in projects if p['name'].lower() == args.name.lower()]
    if not found:
        print(f"Project '{args.name}' not found.")
        return 1
    new_name = args.new_name.strip() if args.new_name else args.name.strip()
    if new_name.lower() == args.name.lower():
        print("New name is same as old name.")
        return 0

    wb = load_workbook(WORKLOG_FILE)
    ws_pr = wb['Projects']
    ws_kp = wb['KPIs']
    ws_st = wb['SubTasks']
    ws_te = wb['Time Entries']
    r = found[0]['row']
    ws_pr.cell(r, 2).value = new_name

    for r2 in range(2, ws_kp.max_row + 1):
        pv = ws_kp.cell(r2, 3).value
        if pv and str(pv).strip().lower() == args.name.lower():
            ws_kp.cell(r2, 3).value = new_name

    for r2 in range(2, ws_st.max_row + 1):
        pv = ws_st.cell(r2, 4).value
        if pv and str(pv).strip().lower() == args.name.lower():
            ws_st.cell(r2, 4).value = new_name

    for r2 in range(2, ws_te.max_row + 1):
        pv = ws_te.cell(r2, 3).value
        if pv and str(pv).strip().lower() == args.name.lower():
            ws_te.cell(r2, 3).value = new_name

    wb.save(WORKLOG_FILE)
    print(f"Renamed project '{args.name}' -> '{new_name}' (updated in Projects, KPIs, SubTasks, Time Entries)")
    gsheets.sync_project_update(args.name, 'name', new_name)
    gsheets.sync_project_rename_propagate(args.name, new_name)
    return 0


def cmd_project_archive(args):
    projects = get_projects()
    found = [p for p in projects if p['name'].lower() == args.name.lower()]
    if not found:
        print(f"Project '{args.name}' not found.")
        return 1
    wb = load_workbook(WORKLOG_FILE)
    ws = wb['Projects']
    r = found[0]['row']
    ws.cell(r, 4).value = 'Archived'
    wb.save(WORKLOG_FILE)
    print(f"Project '{args.name}' archived.")
    gsheets.sync_project_update(args.name, 'status', 'Archived')
    return 0


def countdown(seconds, rest=False):
    """Countdown timer with progress display."""
    bar_width = 30
    start = time.time()
    end = start + seconds
    symbol = ' ' if rest else '\u2588'
    while True:
        remaining = max(0, end - time.time())
        elapsed = seconds - remaining
        pct = elapsed / seconds if seconds > 0 else 1
        filled = int(pct * bar_width)
        bar = symbol * filled + '\u250a' + '\u00b7' * (bar_width - filled)
        mins = int(remaining) // 60
        secs = int(remaining) % 60
        label = 'REST' if rest else 'WORK'
        print(f"\r  {label} [{bar}] {mins:02d}:{secs:02d}  ", end='', flush=True)
        if remaining <= 0:
            print("\a")  # beep
            print()
            return
        time.sleep(0.5)


# ─── Main ────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description='Work Log Timer — CLI tool for WPS Office users',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__)
    sub = parser.add_subparsers(dest='cmd')

    p_start = sub.add_parser('start', help='Start timer')
    p_start.add_argument('-p', '--project', help='Project name (auto: git folder)')
    p_start.add_argument('-t', '--task', help='Father task (auto: git branch)')
    p_start.add_argument('-s', '--subtask', help='Sub task')
    p_start.add_argument('--subtask-code', help='Sub task code')
    p_start.add_argument('-c', '--category', help='Category (default: Development)')
    p_start.add_argument('-d', '--description', help='Description (auto: git commit msg)')

    p_stop = sub.add_parser('stop', help='Stop timer and save')
    p_stop.add_argument('-p', '--project', help='Override project')
    p_stop.add_argument('-t', '--task', help='Override father task')
    p_stop.add_argument('-s', '--subtask', help='Sub task')
    p_stop.add_argument('--subtask-code', help='Sub task code')
    p_stop.add_argument('-c', '--category', help='Override category')
    p_stop.add_argument('-d', '--description', help='Override description')

    sub.add_parser('cancel', help='Cancel running timer')
    sub.add_parser('status', help='Show timer status')
    p_pause = sub.add_parser('pause', help='Pause timer (break, meeting, lunch)')
    p_pause.add_argument('-r', '--reason', help='Pause reason: meeting, lunch, break, review, context-switch, etc.')
    sub.add_parser('continue', help='Resume paused timer')
    sub.add_parser('webhook-test', help='Send test webhook notification')
    sub.add_parser('today', help='Show today summary')
    sub.add_parser('week', help='Show this week summary')

    p_pomo = sub.add_parser('pomo', help='Pomodoro timer')
    p_pomo.add_argument('--work', type=int, default=25, help='Work minutes (default: 25)')
    p_pomo.add_argument('--rest', type=int, default=5, help='Rest minutes (default: 5)')

    # KPI subcommands
    p_kpi = sub.add_parser('kpi', help='Manage father tasks (KPIs)')
    ksub = p_kpi.add_subparsers(dest='kpi_cmd')
    p_kl = ksub.add_parser('list', help='List all father tasks with status')
    p_kl.set_defaults(kpi_func=cmd_kpi_list)
    p_ka = ksub.add_parser('add', help='Add a new father task')
    p_ka.add_argument('-t', '--name', required=True, help='Father task name')
    p_ka.add_argument('-p', '--project', help='Project name')
    p_ka.add_argument('-d', '--deadline', type=int, default=7, help='Deadline in days (default: 7)')
    p_ka.set_defaults(kpi_func=cmd_kpi_add)
    p_kd = ksub.add_parser('done', help='Mark a father task as done')
    p_kd.add_argument('-t', '--name', required=True, help='Father task name')
    p_kd.set_defaults(kpi_func=cmd_kpi_done)
    p_ke = ksub.add_parser('edit', help='Edit father task project/deadline')
    p_ke.add_argument('-t', '--name', required=True, help='Father task name')
    p_ke.add_argument('-p', '--project', help='New project name')
    p_ke.add_argument('-d', '--deadline', type=int, help='New deadline in days')
    p_ke.set_defaults(kpi_func=cmd_kpi_edit)
    p_kr = ksub.add_parser('rename', help='Rename father task')
    p_kr.add_argument('-t', '--name', required=True, help='Current father task name')
    p_kr.add_argument('-n', '--new-name', help='New father task name')
    p_kr.set_defaults(kpi_func=cmd_kpi_rename)
    p_kdel = ksub.add_parser('delete', help='Delete a father task')
    p_kdel.add_argument('-t', '--name', required=True, help='Father task name to delete')
    p_kdel.set_defaults(kpi_func=cmd_kpi_delete)
    p_ks = ksub.add_parser('status', help='Show overall task progress')
    p_ks.set_defaults(kpi_func=cmd_kpi_status)

    # Project subcommands
    p_proj = sub.add_parser('project', help='Manage projects')
    projsub = p_proj.add_subparsers(dest='proj_cmd')
    p_pa = projsub.add_parser('add', help='Add a new project')
    p_pa.add_argument('-n', '--name', required=True, help='Project name')
    p_pa.add_argument('-d', '--description', help='Project description')
    p_pa.set_defaults(proj_func=cmd_project_add)
    p_pl = projsub.add_parser('list', help='List all projects')
    p_pl.set_defaults(proj_func=cmd_project_list)
    p_pr = projsub.add_parser('rename', help='Rename a project')
    p_pr.add_argument('-n', '--name', required=True, help='Current project name')
    p_pr.add_argument('--new-name', help='New project name')
    p_pr.set_defaults(proj_func=cmd_project_rename)
    p_parch = projsub.add_parser('archive', help='Archive a project')
    p_parch.add_argument('-n', '--name', required=True, help='Project name to archive')
    p_parch.set_defaults(proj_func=cmd_project_archive)

    # Subtask subcommands
    p_sub = sub.add_parser('subtask', help='Manage subtasks')
    ssub = p_sub.add_subparsers(dest='sub_cmd')
    p_sa = ssub.add_parser('add', help='Add a new subtask to SubTasks sheet')
    p_sa.add_argument('-s', '--subtask', required=True, help='Subtask name')
    p_sa.add_argument('-t', '--task', required=True, help='Father task name')
    p_sa.add_argument('-p', '--project', help='Project name')
    p_sa.add_argument('--notes', help='Notes')
    p_sa.set_defaults(sub_func=cmd_sub_add)
    p_sl = ssub.add_parser('list', help='List all subtask names from Time Entries')
    p_sl.add_argument('-t', '--task', help='Filter by father task')
    p_sl.set_defaults(sub_func=cmd_sub_list)
    p_se = ssub.add_parser('edit', help='Edit subtask in SubTasks sheet')
    p_se.add_argument('-s', '--subtask', required=True, help='Current subtask name')
    p_se.add_argument('-n', '--new-name', help='New subtask name')
    p_se.add_argument('-t', '--task', help='New father task')
    p_se.add_argument('-p', '--project', help='New project')
    p_se.add_argument('--notes', help='New notes')
    p_se.set_defaults(sub_func=cmd_sub_edit)
    p_sr = ssub.add_parser('rename', help='Rename subtask in all entries')
    p_sr.add_argument('-s', '--subtask', required=True, help='Current subtask name')
    p_sr.add_argument('-n', '--new-name', required=True, help='New subtask name')
    p_sr.set_defaults(sub_func=cmd_sub_rename)
    p_sdel = ssub.add_parser('delete', help='Clear subtask from all matching entries')
    p_sdel.add_argument('-s', '--subtask', required=True, help='Subtask name to clear')
    p_sdel.set_defaults(sub_func=cmd_sub_delete)
    p_sdone = ssub.add_parser('done', help='Mark a subtask as completed')
    p_sdone.add_argument('-s', '--subtask', required=True, help='Subtask name')
    p_sdone.set_defaults(sub_func=cmd_sub_done)
    p_sst = ssub.add_parser('status', help='Show subtask progress')
    p_sst.set_defaults(sub_func=cmd_sub_status)
    p_sstart = ssub.add_parser('start', help='Start timer for an existing subtask')
    p_sstart.add_argument('-s', '--subtask', required=True, help='Subtask name or code')
    p_sstart.add_argument('-p', '--project', help='Project name override')
    p_sstart.add_argument('-t', '--task', help='Father task override')
    p_sstart.add_argument('-c', '--category', help='Category override')
    p_sstart.add_argument('-d', '--description', help='Description override')
    p_sstart.set_defaults(sub_func=cmd_sub_start)

    p_tasks = sub.add_parser('tasks', help='Show tasks grouped by father task')
    p_tasks.add_argument('-t', '--task', help='Filter by father task name')

    args = parser.parse_args()

    if not args.cmd:
        parser.print_help()
        return 1

    cmds = {
        'start': cmd_start,
        'stop': cmd_stop,
        'cancel': cmd_cancel,
        'status': cmd_status,
        'pause': cmd_pause,
        'continue': cmd_continue,
        'today': cmd_today,
        'week': cmd_week,
        'pomo': cmd_pomo,
        'kpi': cmd_kpi,
        'project': cmd_project,
        'tasks': cmd_tasks,
        'subtask': cmd_subtask,
        'webhook-test': webhook.send_test,
    }
    return cmds[args.cmd](args)


if __name__ == '__main__':
    sys.exit(main())

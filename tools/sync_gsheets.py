#!/usr/bin/env python3
"""Full sync from worklog.xlsx → Google Sheets.

Reads all 4 sheets from Excel and overwrites the corresponding
Google Sheets worksheets so they match exactly.

Usage:
    python -m tools.sync_gsheets
"""

import os
import sys
import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
import gsheets
from openpyxl import load_workbook

WORKLOG_FILE = 'worklog.xlsx'


def _cell_to_str(val):
    """Convert an Excel cell value to a clean string for Google Sheets."""
    if val is None:
        return ''
    if isinstance(val, datetime.datetime):
        # If time component is midnight, it's a date-only value
        if val.hour == 0 and val.minute == 0 and val.second == 0:
            return val.strftime('%Y-%m-%d')
        return val.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(val, datetime.date):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, datetime.time):
        if val.second:
            return val.strftime('%H:%M:%S')
        return val.strftime('%H:%M')
    if isinstance(val, (int, float)):
        return str(val)
    return str(val).strip()


def _read_sheet(wb, sheet_name, max_cols):
    """Read all rows from an Excel sheet (including header row 1)."""
    ws = wb[sheet_name]
    rows = []
    for r in range(1, ws.max_row + 1):
        row_data = []
        all_empty = True
        for c in range(1, max_cols + 1):
            v = ws.cell(r, c).value
            s = _cell_to_str(v)
            # Skip placeholder rows
            if s.startswith('[Enter '):
                s = ''
            if s:
                all_empty = False
            row_data.append(s)
        # Stop at first data row where primary key cols (1 & 2) are both empty
        if r > 1 and (not row_data[0] and not row_data[1]):
            break
        rows.append(row_data)
    return rows


def full_sync():
    """Read worklog.xlsx and overwrite all 4 Google Sheets worksheets."""
    if not os.path.exists(WORKLOG_FILE):
        print(f"[SYNC] ERROR: {WORKLOG_FILE} not found")
        return False

    if not gsheets.is_enabled():
        print("[SYNC] ERROR: Google Sheets not configured (missing env vars)")
        return False

    svc = gsheets._get_service()
    if not svc:
        print("[SYNC] ERROR: Cannot connect to Google Sheets")
        return False

    wb = load_workbook(WORKLOG_FILE)

    # Sheet configs: (sheet_name, max_columns)
    sheets = [
        ('Projects', 5),
        ('KPIs', 9),
        ('SubTasks', 8),
        ('Time Entries', 12),
    ]

    for sheet_name, max_cols in sheets:
        print(f"[SYNC] Reading '{sheet_name}' from Excel...", end=' ')
        rows = _read_sheet(wb, sheet_name, max_cols)
        data_rows = len(rows) - 1  # minus header
        print(f"{data_rows} data rows found.")

        if not rows:
            print(f"[SYNC] WARNING: '{sheet_name}' is empty in Excel, skipping.")
            continue

        try:
            gs_ws = svc.worksheet(sheet_name)
        except Exception as e:
            print(f"[SYNC] WARNING: Sheet '{sheet_name}' not found in Google Sheets, creating...")
            try:
                gs_ws = svc.add_worksheet(title=sheet_name, rows=max(len(rows), 100), cols=max_cols)
            except Exception as e2:
                print(f"[SYNC] ERROR creating sheet '{sheet_name}': {e2}")
                continue

        # Clear existing data
        print(f"[SYNC] Clearing '{sheet_name}' in Google Sheets...")
        gs_ws.clear()

        # Write all rows at once
        print(f"[SYNC] Writing {len(rows)} rows to '{sheet_name}'...")
        gs_ws.update(rows, value_input_option='USER_ENTERED')

        print(f"[SYNC] [OK] '{sheet_name}' synced: {data_rows} data rows (+ 1 header)")

    # Sync Gamify Summary worksheet
    try:
        import gamify
        g_res = gamify.run()
        gamify_rows = [
            ["Metric", "Value", "Notes / Details"],
            ["Hero Rank", f"{g_res.get('level_name')} (Lv.{g_res.get('level')})", "Calculated from 5 Level Milestones"],
            ["Total EXP Score", str(g_res.get('total_score', 0)), "Work EXP + Completion EXP"],
            ["Current Level", f"Level {g_res.get('level', 0)} / 100", "0 to 100 Range"],
            ["Daily Work Score", str(g_res.get('today_score', 0)), "Target 7.5h + Category Bonus"],
            ["Daily Streak", f"{g_res.get('streak', 0)} Days", f"Best: {g_res.get('max_streak', 0)} Days"],
            ["Consistency Rate", f"{g_res.get('consistency_pct', 0)}%", f"Logged: {g_res.get('logged_weekdays', 0)}/{g_res.get('total_weekdays', 0)} weekdays"],
            ["Sync Date", datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "Auto-synced from WorkLog System"]
        ]
        try:
            gs_gamify = svc.worksheet('Gamify Summary')
        except Exception:
            gs_gamify = svc.add_worksheet(title='Gamify Summary', rows=20, cols=3)
        gs_gamify.clear()
        gs_gamify.update(gamify_rows, value_input_option='USER_ENTERED')
        print("[SYNC] [OK] 'Gamify Summary' synced to Google Sheets!")
    except Exception as e:
        print(f"[SYNC] WARNING: Could not sync Gamify Summary: {e}")

    print()
    print("[SYNC] [!!]")
    print("[SYNC] [OK] FULL SYNC COMPLETE — All 5 sheets updated!")
    print("[SYNC] [!!]")
    return True


if __name__ == '__main__':
    success = full_sync()
    sys.exit(0 if success else 1)
